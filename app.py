import hashlib
import json
import mimetypes
import os
import re
import secrets
import sqlite3
import string
import subprocess
from dataclasses import dataclass
from datetime import datetime, timedelta
from functools import lru_cache
from http.cookies import SimpleCookie
from io import BytesIO
from pathlib import Path
from threading import Lock
from typing import Any
from urllib.parse import parse_qs, quote
from urllib.error import HTTPError, URLError
from urllib.request import Request as UrlRequest, urlopen
from wsgiref.simple_server import make_server

from jinja2 import Environment, FileSystemLoader, select_autoescape
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

try:
    import psycopg
    from psycopg.rows import dict_row
    from psycopg_pool import ConnectionPool
except ImportError:
    psycopg = None
    dict_row = None
    ConnectionPool = None

BASE_DIR = Path(__file__).resolve().parent
TEMPLATE_DIR = BASE_DIR / "templates"
STATIC_DIR = BASE_DIR / "static"
DATA_DIR = BASE_DIR / "data"
DB_PATH = DATA_DIR / "afe.db"

SESSION_COOKIE = "afe_session"
SESSION_HOURS = 12
PASSWORD_ITERATIONS = 120_000

ROUTES: list[tuple[str, re.Pattern[str], Any]] = []
POSTGRES_POOL: Any | None = None
POSTGRES_POOL_LOCK = Lock()
APP_READY = False
APP_READY_LOCK = Lock()
REFERENCE_API_BASE_URL = os.environ.get("ATME_REFERENCE_API_URL", "https://api.xn--vj4b68z.com").rstrip("/")
INVALID_LOGIN_MESSAGE = "잘못된 로그인 시도입니다. ID나 PW를 다시 확인하시기 바랍니다."

SCHOOL_LEVEL_OPTIONS = ["중학교", "고등학교"]
SEMESTER_OPTIONS = ["1학기", "2학기"]
PROGRAM_STATUS_LABELS = {
    "collecting": "학생 입력 진행 중",
    "teacher_submitted": "강사 제출 완료",
    "completed": "관리자 마감 완료",
}
SUBMISSION_STATUS_LABELS = {
    "student_submitted": "학생 제출",
    "reviewed": "강사 검토 완료",
    "admin_updated": "관리자 검토 완료",
}

PDF_TEMPLATE_PRESETS = [
    {
        "name": "실험 탐구활동 결과질문지",
        "description": "화학실험 활동기록보고서와 연계되는 실험 탐구형 학생 질문지입니다.",
        "report_reference": "화학실험 활동기록보고서.pdf",
        "report_summary": "실험 보고서형 문장으로 정리될 수 있도록 교과 연계, 핵심 키워드, 탐구 확장 계획까지 단계적으로 입력받습니다.",
        "sections": [
            {
                "title": "교과 연계 및 진로 정보",
                "description": "보고서 상단에 들어갈 교과, 단원, 진로 기본 정보를 먼저 정리합니다.",
                "fields": [
                    {"id": "linked_subject_1", "label": "연계 교과 1", "type": "text", "placeholder": "예: 화학실험"},
                    {"id": "linked_unit_1", "label": "연계 단원 1", "type": "text", "placeholder": "예: 물질의 성질 탐구"},
                    {"id": "linked_subject_2", "label": "연계 교과 2", "type": "text", "placeholder": "예: 과학과제연구"},
                    {"id": "linked_unit_2", "label": "연계 단원 2", "type": "text", "placeholder": "예: 연구 설계"},
                    {"id": "career_goal", "label": "희망 진로", "type": "text", "placeholder": "예: 화학 공정 설계 및 시스템 구축 전문가"},
                    {"id": "career_major", "label": "희망 학과 또는 전공", "type": "text", "placeholder": "예: 화학공학과"},
                    {"id": "keywords", "label": "핵심 키워드", "type": "text", "placeholder": "예: 편광, 반응 속도, chirality, 광학 활성"},
                ],
            },
            {
                "title": "탐구 내용 정리",
                "description": "학생이 실제로 작성하는 실험 탐구 서술형 질문입니다.",
                "fields": [
                    {
                        "id": "curriculum_connection",
                        "label": "어떤 교과과정과 기초지식을 활용하여 탐구(실험)활동으로 연계하였습니까?",
                        "type": "textarea",
                        "rows": 5,
                        "placeholder": "교과와 단원을 어떻게 연결했고, 어떤 계기로 탐구를 시작했는지 적어 주세요.",
                    },
                    {
                        "id": "interest_focus",
                        "label": "탐구 과정에서 가장 큰 관심을 가진 개념 또는 활동은 무엇입니까?",
                        "type": "textarea",
                        "rows": 4,
                        "placeholder": "가장 흥미로웠던 개념, 실험 변수, 분석 포인트를 적어 주세요.",
                    },
                    {
                        "id": "research_process",
                        "label": "추가로 조사하거나 익힌 실험 기법, 분석 도구, 데이터 처리 과정은 무엇입니까?",
                        "type": "textarea",
                        "rows": 5,
                        "placeholder": "직접 익힌 기기 사용법, 추가 조사한 내용, 데이터 분석 과정을 구체적으로 적어 주세요.",
                    },
                    {
                        "id": "career_extension",
                        "label": "희망 진로와 연계하여 앞으로 확장하고 싶은 후속 탐구는 무엇입니까?",
                        "type": "textarea",
                        "rows": 4,
                        "placeholder": "앞으로 어떤 실험, 연구, 전공 탐구로 확장하고 싶은지 적어 주세요.",
                    },
                    {
                        "id": "highlight_point",
                        "label": "이번 탐구에서 특히 강조하고 싶은 성과나 강점은 무엇입니까?",
                        "type": "textarea",
                        "rows": 4,
                        "placeholder": "본인이 잘했다고 생각하는 점, 분석의 강점, 의미 있는 성과를 적어 주세요.",
                    },
                    {
                        "id": "future_plan",
                        "label": "이번 활동을 바탕으로 앞으로 어떤 역량과 방향을 키워나가고 싶습니까?",
                        "type": "textarea",
                        "rows": 4,
                        "placeholder": "후속 연구, 진로 확장, 장기적인 성장 계획을 적어 주세요.",
                    },
                ],
            },
        ],
    },
    {
        "name": "사회 탐구활동 결과질문지",
        "description": "미래산업 활동기록보고서와 연계되는 사회 탐구형 학생 질문지입니다.",
        "report_reference": "미래산업 활동기록보고서.pdf",
        "report_summary": "매체, 사회 문제, 미래산업 등 융합 탐구 결과를 보고서형 문장으로 정리할 수 있도록 설계된 템플릿입니다.",
        "sections": [
            {
                "title": "교과 연계 및 진로 정보",
                "description": "학생의 탐구 배경과 진로 방향을 먼저 정리합니다.",
                "fields": [
                    {"id": "linked_subject_1", "label": "연계 교과 1", "type": "text", "placeholder": "예: 언어와 매체"},
                    {"id": "linked_unit_1", "label": "연계 단원 1", "type": "text", "placeholder": "예: 매체의 표현과 활용"},
                    {"id": "linked_subject_2", "label": "연계 교과 2", "type": "text", "placeholder": "예: 사회문제탐구"},
                    {"id": "linked_unit_2", "label": "연계 단원 2", "type": "text", "placeholder": "예: 인공지능의 발전에 따른 예측과 대응 방안"},
                    {"id": "career_goal", "label": "희망 진로", "type": "text", "placeholder": "예: 통역사"},
                    {"id": "career_major", "label": "희망 학과 또는 전공", "type": "text", "placeholder": "예: 통번역학과"},
                    {"id": "keywords", "label": "핵심 키워드", "type": "text", "placeholder": "예: 미래산업, UX, 디자인, 검색엔진"},
                ],
            },
            {
                "title": "탐구 내용 정리",
                "description": "사회 탐구 결과를 보고서 문장으로 풀어낼 수 있도록 돕는 질문입니다.",
                "fields": [
                    {
                        "id": "curriculum_connection",
                        "label": "어떤 교과과정과 기초지식을 활용하여 탐구활동으로 연계하였습니까?",
                        "type": "textarea",
                        "rows": 5,
                        "placeholder": "교과 연계, 탐구 계기, 문제의식을 함께 적어 주세요.",
                    },
                    {
                        "id": "interest_focus",
                        "label": "탐구 과정에서 가장 깊은 관심을 가진 개념 또는 활동은 무엇입니까?",
                        "type": "textarea",
                        "rows": 4,
                        "placeholder": "특히 주목한 개념이나 사례, 사용자 관점 등을 적어 주세요.",
                    },
                    {
                        "id": "research_process",
                        "label": "구체적으로 조사하거나 분석한 사례, 자료, 비교 내용은 무엇입니까?",
                        "type": "textarea",
                        "rows": 5,
                        "placeholder": "기업 사례, 사회 현상, 문화 비교, 자료 조사 내용을 구체적으로 적어 주세요.",
                    },
                    {
                        "id": "career_extension",
                        "label": "희망 진로와 연계하여 앞으로 확장하고 싶은 실천적 탐구는 무엇입니까?",
                        "type": "textarea",
                        "rows": 4,
                        "placeholder": "앞으로 더 탐구하고 싶은 주제나 실제 적용 계획을 적어 주세요.",
                    },
                    {
                        "id": "highlight_point",
                        "label": "이번 탐구에서 특히 강조하고 싶은 관점이나 성과는 무엇입니까?",
                        "type": "textarea",
                        "rows": 4,
                        "placeholder": "탐구를 통해 얻은 통찰, 인간과 기술의 역할, 본인의 강점을 적어 주세요.",
                    },
                    {
                        "id": "future_plan",
                        "label": "이번 활동을 바탕으로 앞으로 어떤 전문 역량을 키워나가고 싶습니까?",
                        "type": "textarea",
                        "rows": 4,
                        "placeholder": "후속 탐구 계획과 진로 역량 성장 방향을 적어 주세요.",
                    },
                ],
            },
        ],
    },
]

EVALUATION_EXAMPLE_PROMPT = """
그룹 활동 보고서 개별 평가 프롬프트

고등학교 교과 세특 작성해 줘.
[작성 지침]
1. 금지 사항: 제공된 내용 외의 새로운 사실 생성 금지.
2. 서술 방식: 관찰자 시점에서 주관적 평가와 불필요한 수식어를 배제하고 객관적 사실을 서술.
3. 용어: 학교생활기록부 기재 요령에 적합한 용어 사용. ('논문' 언급이나 논문명은 '관련 자료' 등으로 대체.)
4. 동사: 수동적 동사(이해함, 정리함, 파악함, 학습함, 확인함 등)를 배제하고 주도적 동사(구조화함, 분석함, 도출함 등)를 사용. 단, 과장된 동사(규명함, 논증함, 정립함 등)는 금지.
5. 문장 연결: '또한', '아울러', '이어서', '나아가', '향후' 등의 적절한 접속어와 지시어를 반드시 활용하여 문장들을 유기적으로 연결.
[문단 구조 및 글자 수: 한 문단으로 출력할 것]
* 도입부(총 3문장): 교과 및 진로 연계를 탐구 초록 형식으로 서술. 첫 문장은 원문에 제시된 탐구 동기를 서술. 이어지는 2문장은 '[교과명]에서 학습한~' 또는 '[교과명]의~' 형태로 구체적인 교과 개념이나 원리가 탐구에 적용된 과정을 서술. 단원명 언급 금지.
* 전개부(총 4문장): 심화 탐구. 탐구 과정에서 활동한 내용을 구체적으로 풀어서 서술.
* 결론부(총 2문장): 후속 연구 및 종합 평가. 원문에 기재된 후속 연구계획이 있으면 미래 시제로 2문장 서술. 후속 연구계획이 기재되지 않았으면 탐구에 대한 종합 평가를 서술.

출력 규칙:
- 반드시 한국어 한 문단으로만 작성.
- 마크다운, 번호, 제목, 따옴표 없이 본문만 출력.
- 학생이 작성한 정보와 프로그램 정보만 사용.
""".strip()

env = Environment(
    loader=FileSystemLoader(str(TEMPLATE_DIR)),
    autoescape=select_autoescape(["html", "xml"]),
    trim_blocks=True,
    lstrip_blocks=True,
)


@dataclass
class Response:
    body: bytes
    status: str = "200 OK"
    headers: list[tuple[str, str]] | None = None

    def as_wsgi(self) -> tuple[str, list[tuple[str, str]], list[bytes]]:
        headers = self.headers[:] if self.headers else []
        if not any(name.lower() == "content-length" for name, _ in headers):
            headers.append(("Content-Length", str(len(self.body))))
        return self.status, headers, [self.body]


class Request:
    def __init__(self, environ: dict[str, Any], db: Any):
        self.environ = environ
        self.db = db
        self.method = environ.get("REQUEST_METHOD", "GET").upper()
        self.path = environ.get("PATH_INFO", "/") or "/"
        self.query = self._parse_query(environ.get("QUERY_STRING", ""))
        self.form = self._parse_form()
        self.cookies = SimpleCookie(environ.get("HTTP_COOKIE", ""))
        session_cookie = self.cookies.get(SESSION_COOKIE)
        self.session = load_session(db, session_cookie.value if session_cookie else None)
        self.flash = pop_flash(db, self.session["id"]) if self.session else None

    @staticmethod
    def _parse_query(raw_query: str) -> dict[str, str]:
        parsed = parse_qs(raw_query, keep_blank_values=True)
        return {key: values[-1] if values else "" for key, values in parsed.items()}

    def _parse_form(self) -> dict[str, str]:
        if self.method != "POST":
            return {}
        try:
            length = int(self.environ.get("CONTENT_LENGTH") or 0)
        except ValueError:
            length = 0
        if length <= 0:
            return {}
        raw = self.environ["wsgi.input"].read(length).decode("utf-8")
        parsed = parse_qs(raw, keep_blank_values=True)
        return {key: values[-1] if values else "" for key, values in parsed.items()}


def route(method: str, pattern: str):
    regex = re.compile(f"^{pattern}$")

    def decorator(func):
        ROUTES.append((method.upper(), regex, func))
        return func

    return decorator


def now() -> datetime:
    return datetime.now().replace(microsecond=0)


def now_iso() -> str:
    return now().isoformat()


def iso_after_hours(hours: int) -> str:
    return (now() + timedelta(hours=hours)).isoformat()


def configured_database_url() -> str:
    return (
        os.environ.get("DATABASE_URL", "").strip()
        or os.environ.get("SUPABASE_DB_URL", "").strip()
    )


def split_sql_script(script: str) -> list[str]:
    return [statement.strip() for statement in script.split(";") if statement.strip()]


def adapt_sql_for_postgres(sql: str) -> str:
    adapted = sql
    adapted = adapted.replace(
        "INTEGER PRIMARY KEY AUTOINCREMENT",
        "INTEGER GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY",
    )
    if "INSERT OR IGNORE INTO" in adapted:
        adapted = adapted.replace("INSERT OR IGNORE INTO", "INSERT INTO")
        stripped = adapted.rstrip()
        suffix = ";" if stripped.endswith(";") else ""
        if suffix:
            stripped = stripped[:-1].rstrip()
        adapted = f"{stripped} ON CONFLICT DO NOTHING{suffix}"
    return adapted.replace("?", "%s")


class DBCursor:
    def __init__(self, cursor: Any):
        self.cursor = cursor

    @property
    def lastrowid(self) -> Any:
        return getattr(self.cursor, "lastrowid", None)

    def fetchone(self) -> Any:
        return self.cursor.fetchone()

    def fetchall(self) -> list[Any]:
        return self.cursor.fetchall()


class DBConnection:
    def __init__(self, raw: Any, dialect: str, releaser: Any | None = None):
        self.raw = raw
        self.dialect = dialect
        self.releaser = releaser

    def execute(self, sql: str, params: tuple[Any, ...] | list[Any] = ()) -> DBCursor:
        if self.dialect == "postgres":
            cursor = self.raw.cursor()
            cursor.execute(adapt_sql_for_postgres(sql), params)
            return DBCursor(cursor)
        return DBCursor(self.raw.execute(sql, params))

    def executescript(self, script: str) -> None:
        for statement in split_sql_script(script):
            self.execute(statement)

    def commit(self) -> None:
        self.raw.commit()

    def rollback(self) -> None:
        self.raw.rollback()

    def close(self) -> None:
        if self.dialect == "postgres":
            try:
                self.raw.rollback()
            except Exception:
                pass
        if self.releaser:
            self.releaser(self.raw)
            return
        self.raw.close()

    def column_names(self, table_name: str) -> list[str]:
        if self.dialect == "postgres":
            rows = self.execute(
                """
                SELECT column_name
                FROM information_schema.columns
                WHERE table_schema = 'public' AND table_name = ?
                ORDER BY ordinal_position
                """,
                (table_name,),
            ).fetchall()
            return [row["column_name"] for row in rows]
        rows = self.execute(f"PRAGMA table_info({table_name})").fetchall()
        return [row["name"] for row in rows]

    def __enter__(self) -> "DBConnection":
        return self

    def __exit__(self, exc_type, exc, tb) -> None:
        try:
            if exc_type:
                self.rollback()
            else:
                self.commit()
        finally:
            self.close()


def connect_db() -> DBConnection:
    database_url = configured_database_url()
    if database_url:
        if psycopg is None or dict_row is None or ConnectionPool is None:
            raise RuntimeError("psycopg가 설치되지 않아 Postgres/Supabase에 연결할 수 없습니다.")
        pool = get_postgres_pool()
        conn = pool.getconn()
        return DBConnection(conn, "postgres", releaser=pool.putconn)

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return DBConnection(conn, "sqlite")


def get_postgres_pool() -> Any:
    global POSTGRES_POOL
    if POSTGRES_POOL is not None:
        return POSTGRES_POOL

    with POSTGRES_POOL_LOCK:
        if POSTGRES_POOL is not None:
            return POSTGRES_POOL
        database_url = configured_database_url()
        if not database_url:
            raise RuntimeError("DATABASE_URL is not configured.")
        min_size = max(1, int(os.environ.get("DB_POOL_MIN", "1")))
        max_size = max(min_size, int(os.environ.get("DB_POOL_MAX", "5")))
        timeout = int(os.environ.get("DB_POOL_TIMEOUT", "15"))
        POSTGRES_POOL = ConnectionPool(
            conninfo=database_url,
            min_size=min_size,
            max_size=max_size,
            timeout=timeout,
            kwargs={"row_factory": dict_row},
            open=True,
        )
        try:
            POSTGRES_POOL.wait()
        except Exception:
            POSTGRES_POOL.close()
            POSTGRES_POOL = None
            raise
        return POSTGRES_POOL


def ensure_runtime_ready() -> None:
    global APP_READY
    if APP_READY:
        return

    with APP_READY_LOCK:
        if APP_READY:
            return
        init_db()
        APP_READY = True


def password_hash(password: str, salt: str | None = None) -> str:
    salt = salt or secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac(
        "sha256", password.encode("utf-8"), salt.encode("utf-8"), PASSWORD_ITERATIONS
    )
    return f"{salt}${digest.hex()}"


def verify_password(password: str, stored_hash: str) -> bool:
    try:
        salt, digest = stored_hash.split("$", 1)
    except ValueError:
        return False
    candidate = password_hash(password, salt).split("$", 1)[1]
    return secrets.compare_digest(candidate, digest)


def json_dump(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False)


def parse_json(value: str | None, fallback: Any) -> Any:
    if not value:
        return fallback
    try:
        return json.loads(value)
    except json.JSONDecodeError:
        return fallback


def parse_questions_from_text(raw_text: str) -> list[str]:
    return [line.strip() for line in raw_text.splitlines() if line.strip()]


ADMIN_PANEL_META: dict[str, dict[str, str]] = {
    "camp-create": {
        "page_title": "캠프 개설",
        "page_description": "교급, 학교, 학년·학기, 담당 강사와 프로그램 유형을 묶어 새 캠프를 개설합니다.",
        "active_key": "camp-create",
    },
    "camp-list": {
        "page_title": "캠프 리스트",
        "page_description": "개설된 캠프 현황을 표로 확인하고 상세, 삭제, Excel 다운로드까지 이어서 관리합니다.",
        "active_key": "camp-list",
    },
    "teacher-create": {
        "page_title": "강사 등록",
        "page_description": "강사 기본 정보와 계정, 코드명, 메모를 등록해 운영 준비를 마칩니다.",
        "active_key": "teacher-create",
    },
    "teacher-list": {
        "page_title": "강사 리스트",
        "page_description": "등록된 강사를 확인하고 캠프 배정과 배정 취소를 한 곳에서 처리합니다.",
        "active_key": "teacher-list",
    },
    "teacher-irregular": {
        "page_title": "강사 배정현황",
        "page_description": "강사별 프로그램 배정 상태와 학생 제출 현황을 운영용 표로 확인합니다.",
        "active_key": "teacher-irregular",
    },
    "template-create": {
        "page_title": "유형 입력",
        "page_description": "새 프로그램 유형을 만들고 기본 질문과 평가 프롬프트를 저장합니다.",
        "active_key": "template-create",
    },
    "template-manage": {
        "page_title": "유형 관리",
        "page_description": "저장된 유형을 수정하고 사용 현황을 확인하며 운영 중인 유형을 정리합니다.",
        "active_key": "template-manage",
    },
    "prompt-manage": {
        "page_title": "프롬프트 관리",
        "page_description": "강사 생활기록부 생성에 사용하는 GPT 프롬프트를 한 곳에서 관리합니다.",
        "active_key": "prompt-manage",
    },
}


def current_admin_panel(request: Request) -> str:
    panel = request.query.get("panel", "camp-create").strip()
    if panel not in ADMIN_PANEL_META:
        return "camp-create"
    return panel


def admin_panel_path(panel: str = "camp-create") -> str:
    resolved_panel = panel if panel in ADMIN_PANEL_META else "camp-create"
    return f"/admin?panel={resolved_panel}"


def get_meta_value(db: sqlite3.Connection, key: str, default: str = "") -> str:
    row = db.execute("SELECT value FROM app_meta WHERE key = ?", (key,)).fetchone()
    if not row:
        return default
    return str(row["value"] or default)


def set_meta_value(db: sqlite3.Connection, key: str, value: str) -> None:
    db.execute(
        """
        INSERT INTO app_meta (key, value) VALUES (?, ?)
        ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value
        """,
        (key, value),
    )
    db.commit()


def get_teacher_generation_prompt(
    db: sqlite3.Connection,
    program: sqlite3.Row | dict[str, Any] | None = None,
) -> str:
    global_prompt = get_meta_value(db, "teacher_generation_prompt", "").strip()
    if global_prompt:
        return global_prompt
    if program is not None:
        program_prompt = (program.get("prompt_text") if isinstance(program, dict) else program["prompt_text"]) or ""
        if str(program_prompt).strip():
            return str(program_prompt).strip()
    return EVALUATION_EXAMPLE_PROMPT.strip()


@lru_cache(maxsize=256)
def fetch_reference_api_payload(path: str) -> dict[str, Any]:
    url = f"{REFERENCE_API_BASE_URL}{path}"
    upstream_request = UrlRequest(
        url,
        headers={
            "Accept": "application/json",
            "User-Agent": "AFE-Camp-Manager/1.0",
        },
    )
    try:
        with urlopen(upstream_request, timeout=12) as response:
            body = response.read().decode("utf-8")
    except URLError:
        body = ""
        last_error: Exception | None = None
        for command in (["curl", "--silent", "--show-error", "--fail", "--max-time", "12", url], ["curl.exe", "--silent", "--show-error", "--fail", "--max-time", "12", url]):
            try:
                completed = subprocess.run(
                    command,
                    capture_output=True,
                    check=True,
                    text=True,
                    encoding="utf-8",
                    timeout=15,
                )
                body = completed.stdout
                break
            except FileNotFoundError as exc:
                last_error = exc
            except subprocess.SubprocessError as exc:
                last_error = exc
        if not body:
            raise URLError(last_error or "Reference API connection failed")
    payload = parse_json(body, {})
    if not isinstance(payload, dict):
        raise ValueError("Invalid reference API response")
    return payload


def normalize_template_schema(raw_schema: Any, fallback_title: str = "") -> dict[str, Any]:
    if isinstance(raw_schema, dict) and isinstance(raw_schema.get("sections"), list):
        sections = []
        flat_fields = []
        field_counter = 0
        for section_index, section in enumerate(raw_schema["sections"], start=1):
            section_title = section.get("title") or f"섹션 {section_index}"
            normalized_section = {
                "title": section_title,
                "description": section.get("description", ""),
                "fields": [],
            }
            for field in section.get("fields", []):
                field_counter += 1
                field_id = field.get("id") or f"field_{field_counter}"
                field_type = field.get("type", "textarea")
                normalized_field = {
                    "id": field_id,
                    "label": field.get("label") or f"문항 {field_counter}",
                    "type": field_type if field_type in {"text", "textarea"} else "textarea",
                    "placeholder": field.get("placeholder", ""),
                    "rows": int(field.get("rows", 4)),
                    "required": field.get("required", True),
                    "hint": field.get("hint", ""),
                    "section_title": section_title,
                }
                normalized_section["fields"].append(normalized_field)
                flat_fields.append(normalized_field)
            sections.append(normalized_section)
        return {
            "title": raw_schema.get("title") or fallback_title,
            "description": raw_schema.get("description", ""),
            "report_reference": raw_schema.get("report_reference", ""),
            "report_summary": raw_schema.get("report_summary", ""),
            "sections": sections,
            "flat_fields": flat_fields,
        }

    questions = raw_schema if isinstance(raw_schema, list) else []
    flat_fields = []
    for index, question in enumerate(questions, start=1):
        if isinstance(question, dict) and question.get("label"):
            flat_fields.append(
                {
                    "id": question.get("id") or f"field_{index}",
                    "label": question["label"],
                    "type": question.get("type", "textarea"),
                    "placeholder": question.get("placeholder", ""),
                    "rows": int(question.get("rows", 4)),
                    "required": question.get("required", True),
                    "hint": question.get("hint", ""),
                    "section_title": "기본 질문",
                }
            )
        elif isinstance(question, str):
            flat_fields.append(
                {
                    "id": f"answer_{index - 1}",
                    "label": question,
                    "type": "textarea",
                    "placeholder": "답변을 입력해 주세요.",
                    "rows": 5,
                    "required": True,
                    "hint": "",
                    "section_title": "기본 질문",
                }
            )
    return {
        "title": fallback_title,
        "description": "",
        "report_reference": "",
        "report_summary": "",
        "sections": [
            {
                "title": "기본 질문",
                "description": "",
                "fields": flat_fields,
            }
        ],
        "flat_fields": flat_fields,
    }


def get_template_schema(record: sqlite3.Row | dict[str, Any]) -> dict[str, Any]:
    raw_schema = parse_json(record["questions_json"], [])
    record_keys = set(record.keys()) if hasattr(record, "keys") else set(record)
    if "name" in record_keys:
        fallback_title = record["name"]
    elif "template_name" in record_keys:
        fallback_title = record["template_name"]
    else:
        fallback_title = ""
    return normalize_template_schema(raw_schema, fallback_title=fallback_title)


def get_template_card(record: sqlite3.Row | dict[str, Any]) -> dict[str, Any]:
    item = dict(record)
    schema = get_template_schema(item)
    item["schema"] = schema
    item["field_count"] = len(schema["flat_fields"])
    item["preview_fields"] = [field["label"] for field in schema["flat_fields"][:4]]
    item["question_lines"] = "\n".join(field["label"] for field in schema["flat_fields"])
    item["prompt_text"] = (item.get("prompt_text") or EVALUATION_EXAMPLE_PROMPT).strip()
    item["prompt_preview"] = item["prompt_text"][:180]
    return item


def status_label(status: str, kind: str = "program") -> str:
    mapping = PROGRAM_STATUS_LABELS if kind == "program" else SUBMISSION_STATUS_LABELS
    return mapping.get(status, status)


def format_datetime(value: str | None) -> str:
    if not value:
        return "-"
    try:
        parsed = datetime.fromisoformat(value)
    except ValueError:
        return value
    return parsed.strftime("%Y-%m-%d %H:%M")


def final_evaluation(submission: sqlite3.Row | dict[str, Any]) -> str:
    admin_feedback = (submission["admin_feedback"] or "").strip()
    if admin_feedback:
        return admin_feedback
    return (submission["teacher_evaluation"] or "").strip()


def init_db(*, skip_bootstrap: bool = False) -> None:
    DATA_DIR.mkdir(exist_ok=True)
    with connect_db() as db:
        db.executescript(
            """
            CREATE TABLE IF NOT EXISTS admins (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT NOT NULL UNIQUE,
                password_hash TEXT NOT NULL,
                display_name TEXT NOT NULL,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS teachers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                email TEXT,
                access_code TEXT NOT NULL UNIQUE,
                university TEXT DEFAULT '',
                school_name TEXT DEFAULT '',
                department_name TEXT DEFAULT '',
                username TEXT DEFAULT '',
                password_hash TEXT DEFAULT '',
                temporary_password TEXT DEFAULT '',
                memo TEXT DEFAULT '',
                academic_info TEXT DEFAULT '',
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS program_templates (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                description TEXT,
                prompt_text TEXT DEFAULT '',
                questions_json TEXT NOT NULL,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS programs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                school_name TEXT NOT NULL,
                school_level TEXT NOT NULL,
                year INTEGER NOT NULL,
                semester TEXT NOT NULL,
                template_id INTEGER,
                template_name TEXT NOT NULL,
                template_description TEXT,
                prompt_text TEXT DEFAULT '',
                questions_json TEXT NOT NULL,
                teacher_id INTEGER NOT NULL,
                program_code TEXT NOT NULL UNIQUE,
                status TEXT NOT NULL DEFAULT 'collecting',
                teacher_submitted_at TEXT,
                created_at TEXT NOT NULL,
                FOREIGN KEY (template_id) REFERENCES program_templates(id),
                FOREIGN KEY (teacher_id) REFERENCES teachers(id)
            );

            CREATE TABLE IF NOT EXISTS program_teachers (
                program_id INTEGER NOT NULL,
                teacher_id INTEGER NOT NULL,
                assigned_at TEXT NOT NULL,
                PRIMARY KEY (program_id, teacher_id),
                FOREIGN KEY (program_id) REFERENCES programs(id) ON DELETE CASCADE,
                FOREIGN KEY (teacher_id) REFERENCES teachers(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS submissions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                program_id INTEGER NOT NULL,
                student_number TEXT NOT NULL,
                student_name TEXT NOT NULL,
                desired_major TEXT NOT NULL,
                answers_json TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'student_submitted',
                student_submitted_at TEXT NOT NULL,
                teacher_summary TEXT DEFAULT '',
                teacher_evaluation TEXT DEFAULT '',
                teacher_updated_at TEXT,
                ai_suggestion TEXT DEFAULT '',
                ai_generated_at TEXT,
                ai_model TEXT DEFAULT '',
                admin_feedback TEXT DEFAULT '',
                admin_updated_at TEXT,
                UNIQUE (program_id, student_number),
                FOREIGN KEY (program_id) REFERENCES programs(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS student_drafts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                program_id INTEGER NOT NULL,
                student_name TEXT NOT NULL,
                student_number TEXT NOT NULL,
                desired_major TEXT DEFAULT '',
                answers_json TEXT NOT NULL DEFAULT '{}',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE (program_id, student_name, student_number),
                FOREIGN KEY (program_id) REFERENCES programs(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS sessions (
                id TEXT PRIMARY KEY,
                role TEXT NOT NULL,
                admin_username TEXT,
                teacher_id INTEGER,
                program_id INTEGER,
                context_json TEXT NOT NULL DEFAULT '{}',
                created_at TEXT NOT NULL,
                expires_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS app_meta (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL
            );
            """
        )
        ensure_teacher_schema(db)
        ensure_template_schema_columns(db)
        ensure_program_schema_columns(db)
        ensure_submission_schema(db)
        ensure_student_draft_schema(db)
        ensure_program_teacher_schema(db)
        if not skip_bootstrap:
            ensure_bootstrap_data(db)
            ensure_pdf_template_presets(db)


def ensure_teacher_schema(db: sqlite3.Connection) -> None:
    columns = set(db.column_names("teachers"))
    if "university" not in columns:
        db.execute("ALTER TABLE teachers ADD COLUMN university TEXT DEFAULT ''")
    if "school_name" not in columns:
        db.execute("ALTER TABLE teachers ADD COLUMN school_name TEXT DEFAULT ''")
    if "department_name" not in columns:
        db.execute("ALTER TABLE teachers ADD COLUMN department_name TEXT DEFAULT ''")
    if "username" not in columns:
        db.execute("ALTER TABLE teachers ADD COLUMN username TEXT DEFAULT ''")
    if "password_hash" not in columns:
        db.execute("ALTER TABLE teachers ADD COLUMN password_hash TEXT DEFAULT ''")
    if "temporary_password" not in columns:
        db.execute("ALTER TABLE teachers ADD COLUMN temporary_password TEXT DEFAULT ''")
    if "memo" not in columns:
        db.execute("ALTER TABLE teachers ADD COLUMN memo TEXT DEFAULT ''")
    if "academic_info" not in columns:
        db.execute("ALTER TABLE teachers ADD COLUMN academic_info TEXT DEFAULT ''")
    db.commit()

    rows = db.execute(
        """
        SELECT
            id,
            name,
            username,
            password_hash,
            temporary_password,
            access_code,
            university,
            school_name,
            department_name
        FROM teachers
        """
    ).fetchall()
    for row in rows:
        username = (row["username"] or "").strip()
        temp_password = (row["temporary_password"] or "").strip()
        password_value = (row["password_hash"] or "").strip()
        school_name = (row["school_name"] or "").strip()
        department_name = (row["department_name"] or "").strip()
        changed = False
        if not username:
            username = generate_teacher_username(db, row["name"], teacher_id=row["id"])
            changed = True
        if not temp_password:
            temp_password = generate_teacher_password(row["id"])
            changed = True
        if not password_value:
            password_value = password_hash(temp_password)
            changed = True
        if not school_name and (row["university"] or "").strip():
            school_name = (row["university"] or "").strip()
            changed = True
        if changed:
            db.execute(
                """
                UPDATE teachers
                SET username = ?, temporary_password = ?, password_hash = ?, school_name = ?, department_name = ?
                WHERE id = ?
                """,
                (username, temp_password, password_value, school_name, department_name, row["id"]),
            )
        db.commit()


def ensure_template_schema_columns(db: sqlite3.Connection) -> None:
    columns = set(db.column_names("program_templates"))
    if "prompt_text" not in columns:
        db.execute("ALTER TABLE program_templates ADD COLUMN prompt_text TEXT DEFAULT ''")
        db.commit()

    rows = db.execute("SELECT id, prompt_text FROM program_templates").fetchall()
    for row in rows:
        if not (row["prompt_text"] or "").strip():
            db.execute(
                "UPDATE program_templates SET prompt_text = ? WHERE id = ?",
                (EVALUATION_EXAMPLE_PROMPT, row["id"]),
            )
    db.commit()


def ensure_program_schema_columns(db: sqlite3.Connection) -> None:
    columns = set(db.column_names("programs"))
    if "prompt_text" not in columns:
        db.execute("ALTER TABLE programs ADD COLUMN prompt_text TEXT DEFAULT ''")
        db.commit()

    programs = db.execute("SELECT id, template_id, prompt_text, school_level FROM programs").fetchall()
    for program in programs:
        if program["school_level"] in {"1학년", "2학년", "3학년"}:
            db.execute("UPDATE programs SET school_level = '고등학교' WHERE id = ?", (program["id"],))
        if (program["prompt_text"] or "").strip():
            continue
        prompt_text = EVALUATION_EXAMPLE_PROMPT
        if program["template_id"]:
            template = db.execute(
                "SELECT prompt_text FROM program_templates WHERE id = ?",
                (program["template_id"],),
            ).fetchone()
            if template and (template["prompt_text"] or "").strip():
                prompt_text = template["prompt_text"]
        db.execute("UPDATE programs SET prompt_text = ? WHERE id = ?", (prompt_text, program["id"]))
    db.commit()


def ensure_program_teacher_schema(db: sqlite3.Connection) -> None:
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS program_teachers (
            program_id INTEGER NOT NULL,
            teacher_id INTEGER NOT NULL,
            assigned_at TEXT NOT NULL,
            PRIMARY KEY (program_id, teacher_id),
            FOREIGN KEY (program_id) REFERENCES programs(id) ON DELETE CASCADE,
            FOREIGN KEY (teacher_id) REFERENCES teachers(id) ON DELETE CASCADE
        )
        """
    )
    rows = db.execute(
        "SELECT id, teacher_id, created_at FROM programs WHERE teacher_id IS NOT NULL"
    ).fetchall()
    for row in rows:
        db.execute(
            """
            INSERT OR IGNORE INTO program_teachers (program_id, teacher_id, assigned_at)
            VALUES (?, ?, ?)
            """,
            (row["id"], row["teacher_id"], row["created_at"] or now_iso()),
        )
    db.commit()


def ensure_submission_schema(db: sqlite3.Connection) -> None:
    columns = set(db.column_names("submissions"))
    if "ai_suggestion" not in columns:
        db.execute("ALTER TABLE submissions ADD COLUMN ai_suggestion TEXT DEFAULT ''")
    if "ai_generated_at" not in columns:
        db.execute("ALTER TABLE submissions ADD COLUMN ai_generated_at TEXT")
    if "ai_model" not in columns:
        db.execute("ALTER TABLE submissions ADD COLUMN ai_model TEXT DEFAULT ''")
    db.commit()


def create_student_drafts_table(db: sqlite3.Connection) -> None:
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS student_drafts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            program_id INTEGER NOT NULL,
            student_name TEXT NOT NULL,
            student_number TEXT NOT NULL,
            desired_major TEXT DEFAULT '',
            answers_json TEXT NOT NULL DEFAULT '{}',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            UNIQUE (program_id, student_name, student_number),
            FOREIGN KEY (program_id) REFERENCES programs(id) ON DELETE CASCADE
        )
        """
    )


def ensure_student_draft_schema(db: sqlite3.Connection) -> None:
    columns = set(db.column_names("student_drafts"))
    if not columns:
        create_student_drafts_table(db)
        db.execute(
            "INSERT OR IGNORE INTO app_meta (key, value) VALUES (?, ?)",
            ("student_draft_identity_v2", "1"),
        )
        db.commit()
        return

    migration_flag = db.execute(
        "SELECT value FROM app_meta WHERE key = ?",
        ("student_draft_identity_v2",),
    ).fetchone()
    if migration_flag and migration_flag["value"] == "1":
        return

    db.execute("ALTER TABLE student_drafts RENAME TO student_drafts_legacy")
    create_student_drafts_table(db)
    legacy_rows = db.execute(
        """
        SELECT program_id, student_name, COALESCE(student_number, '') AS student_number,
               COALESCE(desired_major, '') AS desired_major, answers_json, created_at, updated_at
        FROM student_drafts_legacy
        ORDER BY updated_at ASC, created_at ASC
        """
    ).fetchall()
    for row in legacy_rows:
        student_name = (row["student_name"] or "").strip()
        student_number = (row["student_number"] or "").strip()
        if not student_name or not student_number:
            continue
        db.execute(
            """
            INSERT OR IGNORE INTO student_drafts (
                program_id, student_name, student_number, desired_major,
                answers_json, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                row["program_id"],
                student_name,
                student_number,
                row["desired_major"] or "",
                row["answers_json"] or "{}",
                row["created_at"] or now_iso(),
                row["updated_at"] or now_iso(),
            ),
        )
    db.execute("DROP TABLE student_drafts_legacy")
    db.execute(
        """
        INSERT INTO app_meta (key, value) VALUES (?, ?)
        ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value
        """,
        ("student_draft_identity_v2", "1"),
    )
    db.commit()


def ensure_bootstrap_data(db: sqlite3.Connection) -> None:
    bootstrap_done = db.execute(
        "SELECT value FROM app_meta WHERE key = ?",
        ("bootstrap_completed",),
    ).fetchone()
    if bootstrap_done:
        return

    has_existing_data = any(
        db.execute(f"SELECT 1 FROM {table_name} LIMIT 1").fetchone()
        for table_name in ("admins", "teachers", "program_templates", "programs", "submissions")
    )
    if not has_existing_data:
        seed_defaults(db)

    db.execute(
        "INSERT OR REPLACE INTO app_meta (key, value) VALUES (?, ?)",
        ("bootstrap_completed", now_iso()),
    )
    db.commit()


def ensure_pdf_template_presets(db: sqlite3.Connection) -> None:
    imported = db.execute(
        "SELECT value FROM app_meta WHERE key = ?",
        ("pdf_template_presets_v1",),
    ).fetchone()
    if imported:
        return

    for preset in PDF_TEMPLATE_PRESETS:
        exists = db.execute(
            "SELECT 1 FROM program_templates WHERE name = ?",
            (preset["name"],),
        ).fetchone()
        if exists:
            continue
        db.execute(
            """
            INSERT INTO program_templates (name, description, prompt_text, questions_json, created_at)
            VALUES (?, ?, ?, ?, ?)
            """,
            (
                preset["name"],
                preset["description"],
                EVALUATION_EXAMPLE_PROMPT,
                json_dump(preset),
                now_iso(),
            ),
        )

    db.execute(
        "INSERT OR REPLACE INTO app_meta (key, value) VALUES (?, ?)",
        ("pdf_template_presets_v1", now_iso()),
    )
    db.commit()


def seed_defaults(db: sqlite3.Connection) -> None:
    admin_exists = db.execute("SELECT 1 FROM admins LIMIT 1").fetchone()
    if not admin_exists:
        db.execute(
            """
            INSERT INTO admins (username, password_hash, display_name, created_at)
            VALUES (?, ?, ?, ?)
            """,
            ("admin", password_hash("afe1234!"), "AFE 관리자", now_iso()),
        )

    default_templates = [
        {
            "name": "일반 탐구",
            "description": "자기주도 탐구 활동을 바탕으로 작성하는 일반형 산출물 질문지입니다.",
            "questions": [
                "이번 프로젝트에서 다룬 주제와 선택한 이유를 설명해 주세요.",
                "탐구를 진행하면서 가장 중요하게 참고한 자료 또는 근거는 무엇인가요?",
                "이번 활동을 통해 배우거나 새롭게 알게 된 점을 적어 주세요.",
                "강사와의 면담에서 추가로 받고 싶은 피드백이 있다면 적어 주세요.",
            ],
        },
        {
            "name": "실험",
            "description": "가설, 실험 과정, 결과 해석 중심으로 작성하는 실험형 질문지입니다.",
            "questions": [
                "실험의 목적과 가설을 정리해 주세요.",
                "실험 절차와 사용한 도구 또는 재료를 설명해 주세요.",
                "관찰한 결과와 예상과 달랐던 점이 있었다면 적어 주세요.",
                "실험 결과를 바탕으로 어떤 결론을 내렸는지 설명해 주세요.",
            ],
        },
        {
            "name": "데이터",
            "description": "데이터 수집, 분석, 해석 과정을 정리하는 데이터형 질문지입니다.",
            "questions": [
                "수집한 데이터의 종류와 출처를 적어 주세요.",
                "데이터를 어떤 기준으로 정리하거나 전처리했는지 설명해 주세요.",
                "분석 과정에서 발견한 핵심 패턴이나 인사이트를 적어 주세요.",
                "이번 데이터 분석이 진로와 어떤 관련이 있었는지 적어 주세요.",
            ],
        },
    ]
    for template in default_templates:
        exists = db.execute(
            "SELECT 1 FROM program_templates WHERE name = ?",
            (template["name"],),
        ).fetchone()
        if not exists:
            db.execute(
                """
                INSERT INTO program_templates (name, description, prompt_text, questions_json, created_at)
                VALUES (?, ?, ?, ?, ?)
                """,
                (
                    template["name"],
                    template["description"],
                    EVALUATION_EXAMPLE_PROMPT,
                    json_dump(template["questions"]),
                    now_iso(),
                ),
            )

    teacher_row = db.execute("SELECT * FROM teachers WHERE access_code = ?", ("TCHR2026",)).fetchone()
    if not teacher_row:
        default_teacher_password = "AFE!T0001"
        db.execute(
            """
            INSERT INTO teachers (
                name, email, access_code, university, school_name, department_name,
                username, password_hash, temporary_password, memo, academic_info, created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "김강사",
                "teacher@example.com",
                "TCHR2026",
                "KAIST",
                "KAIST",
                "박사과정",
                "teacher001",
                password_hash(default_teacher_password),
                default_teacher_password,
                "샘플 강사 계정",
                "학교 신분\nKAIST 박사과정",
                now_iso(),
            ),
        )
        teacher_row = db.execute("SELECT * FROM teachers WHERE access_code = ?", ("TCHR2026",)).fetchone()

    if not db.execute("SELECT 1 FROM programs LIMIT 1").fetchone():
        template = db.execute(
            "SELECT * FROM program_templates WHERE name = ?",
            ("일반 탐구",),
        ).fetchone()
        db.execute(
            """
            INSERT INTO programs (
                title, school_name, school_level, year, semester,
                template_id, template_name, template_description, prompt_text, questions_json,
                teacher_id, program_code, status, teacher_submitted_at, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "2026 1학기 일반 탐구 시범 프로그램",
                "충주대원고등학교",
                "고등학교",
                2026,
                "1학기",
                template["id"],
                template["name"],
                template["description"],
                template["prompt_text"] or EVALUATION_EXAMPLE_PROMPT,
                template["questions_json"],
                teacher_row["id"],
                "20261001",
                "collecting",
                None,
                now_iso(),
            ),
        )
        program_row = db.execute(
            "SELECT * FROM programs WHERE program_code = ?",
            ("20261001",),
        ).fetchone()
        db.execute(
            """
            INSERT INTO submissions (
                program_id, student_number, student_name, desired_major,
                answers_json, status, student_submitted_at,
                teacher_summary, teacher_evaluation, teacher_updated_at,
                admin_feedback, admin_updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                program_row["id"],
                "30101",
                "홍길동",
                "컴퓨터공학과",
                json_dump(
                    [
                        {
                            "question": "이번 프로젝트에서 다룬 주제와 선택한 이유를 설명해 주세요.",
                            "answer": "AI를 활용한 교육 프로그램 기획 과정을 주제로 삼았습니다.",
                        },
                        {
                            "question": "탐구를 진행하면서 가장 중요하게 참고한 자료 또는 근거는 무엇인가요?",
                            "answer": "학교 수업 자료와 관련 교육 프로그램 운영 사례를 참고했습니다.",
                        },
                        {
                            "question": "이번 활동을 통해 배우거나 새롭게 알게 된 점을 적어 주세요.",
                            "answer": "서비스를 설계할 때 사용자 흐름을 먼저 정리하는 것이 중요하다는 점을 배웠습니다.",
                        },
                        {
                            "question": "강사와의 면담에서 추가로 받고 싶은 피드백이 있다면 적어 주세요.",
                            "answer": "산출물 정리 방식과 발표 구조에 대한 피드백을 받고 싶습니다.",
                        },
                    ]
                ),
                "student_submitted",
                now_iso(),
                "",
                "",
                None,
                "",
                None,
            ),
        )
    db.commit()


def load_session(db: sqlite3.Connection, session_id: str | None) -> dict[str, Any] | None:
    if not session_id:
        return None
    row = db.execute(
        "SELECT * FROM sessions WHERE id = ? AND expires_at > ?",
        (session_id, now_iso()),
    ).fetchone()
    if not row:
        return None
    session = dict(row)
    session["context"] = parse_json(row["context_json"], {})
    return session


def create_session(
    db: sqlite3.Connection,
    *,
    role: str,
    admin_username: str | None = None,
    teacher_id: int | None = None,
    program_id: int | None = None,
    context: dict[str, Any] | None = None,
) -> str:
    session_id = secrets.token_urlsafe(24)
    db.execute(
        """
        INSERT INTO sessions (id, role, admin_username, teacher_id, program_id, context_json, created_at, expires_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            session_id,
            role,
            admin_username,
            teacher_id,
            program_id,
            json_dump(context or {}),
            now_iso(),
            iso_after_hours(SESSION_HOURS),
        ),
    )
    db.commit()
    return session_id


def destroy_session(db: sqlite3.Connection, session_id: str | None) -> None:
    if not session_id:
        return
    db.execute("DELETE FROM sessions WHERE id = ?", (session_id,))
    db.commit()


def set_flash(db: sqlite3.Connection, session_id: str | None, message: str, tone: str = "info") -> None:
    if not session_id:
        return
    row = db.execute("SELECT context_json FROM sessions WHERE id = ?", (session_id,)).fetchone()
    if not row:
        return
    context = parse_json(row["context_json"], {})
    context["_flash"] = {"message": message, "tone": tone}
    db.execute(
        "UPDATE sessions SET context_json = ?, expires_at = ? WHERE id = ?",
        (json_dump(context), iso_after_hours(SESSION_HOURS), session_id),
    )
    db.commit()


def pop_flash(db: sqlite3.Connection, session_id: str) -> dict[str, str] | None:
    row = db.execute("SELECT context_json FROM sessions WHERE id = ?", (session_id,)).fetchone()
    if not row:
        return None
    context = parse_json(row["context_json"], {})
    flash = context.pop("_flash", None)
    db.execute(
        "UPDATE sessions SET context_json = ?, expires_at = ? WHERE id = ?",
        (json_dump(context), iso_after_hours(SESSION_HOURS), session_id),
    )
    db.commit()
    return flash


def update_session_context(
    db: sqlite3.Connection,
    session_id: str | None,
    updates: dict[str, Any],
) -> dict[str, Any]:
    if not session_id:
        return {}
    row = db.execute("SELECT context_json FROM sessions WHERE id = ?", (session_id,)).fetchone()
    if not row:
        return {}
    context = parse_json(row["context_json"], {})
    for key, value in updates.items():
        if value is None:
            context.pop(key, None)
        else:
            context[key] = value
    db.execute(
        "UPDATE sessions SET context_json = ?, expires_at = ? WHERE id = ?",
        (json_dump(context), iso_after_hours(SESSION_HOURS), session_id),
    )
    db.commit()
    return context


def generate_program_code(db: sqlite3.Connection) -> str:
    alphabet = "ABCDEFGHJKLMNPQRSTUVWXYZ23456789"
    while True:
        code = "".join(secrets.choice(alphabet) for _ in range(10))
        exists = db.execute(
            "SELECT 1 FROM programs WHERE program_code = ?",
            (code,),
        ).fetchone()
        if not exists:
            return code


def generate_teacher_code(db: sqlite3.Connection) -> str:
    alphabet = "ABCDEFGHJKLMNPQRSTUVWXYZ23456789"
    while True:
        code = "".join(secrets.choice(alphabet) for _ in range(8))
        exists = db.execute(
            "SELECT 1 FROM teachers WHERE access_code = ?",
            (code,),
        ).fetchone()
        if not exists:
            return code


def generate_teacher_username(db: sqlite3.Connection, name: str, *, teacher_id: int | None = None) -> str:
    base = re.sub(r"[^a-z0-9]", "", name.lower())
    if not base:
        base = "teacher"
    base = base[:10]
    if teacher_id is not None:
        candidate = f"{base}{teacher_id:03d}"[:16]
        exists = db.execute(
            "SELECT 1 FROM teachers WHERE username = ? AND id != ?",
            (candidate, teacher_id),
        ).fetchone()
        if not exists:
            return candidate
    while True:
        suffix = "".join(secrets.choice(string.digits) for _ in range(4))
        candidate = f"{base}{suffix}"[:16]
        exists = db.execute(
            "SELECT 1 FROM teachers WHERE username = ?",
            (candidate,),
        ).fetchone()
        if not exists:
            return candidate


def generate_teacher_password(teacher_id: int | None = None) -> str:
    if teacher_id is not None:
        return f"AFE!T{teacher_id:04d}"
    alphabet = string.ascii_letters + string.digits
    return "AFE!" + "".join(secrets.choice(alphabet) for _ in range(8))


def html_response(html: str, status: str = "200 OK", headers: list[tuple[str, str]] | None = None) -> Response:
    base_headers = [("Content-Type", "text/html; charset=utf-8")]
    if headers:
        base_headers.extend(headers)
    return Response(body=html.encode("utf-8"), status=status, headers=base_headers)


def redirect_response(location: str, headers: list[tuple[str, str]] | None = None) -> Response:
    if location.startswith("/admin/login?error="):
        location = f"/?role=admin&error={quote(INVALID_LOGIN_MESSAGE)}"
    elif location.startswith("/?role=teacher&error="):
        location = f"/?role=teacher&error={quote(INVALID_LOGIN_MESSAGE)}"
    elif location.startswith("/?role=student&error="):
        location = f"/?role=student&error={quote(INVALID_LOGIN_MESSAGE)}"
    base_headers = [("Location", location)]
    if headers:
        base_headers.extend(headers)
    return Response(body=b"", status="302 Found", headers=base_headers)


def login_error_redirect(role: str) -> Response:
    return redirect_response(f"/?role={quote(role)}&error={quote(INVALID_LOGIN_MESSAGE)}")


def text_response(text: str, status: str = "200 OK") -> Response:
    return Response(
        body=text.encode("utf-8"),
        status=status,
        headers=[("Content-Type", "text/plain; charset=utf-8")],
    )


def json_response(payload: Any, status: str = "200 OK") -> Response:
    return Response(
        body=json_dump(payload).encode("utf-8"),
        status=status,
        headers=[
            ("Content-Type", "application/json; charset=utf-8"),
            ("Cache-Control", "public, max-age=3600"),
        ],
    )


def bytes_response(
    content: bytes,
    *,
    content_type: str,
    filename: str | None = None,
) -> Response:
    headers = [("Content-Type", content_type)]
    if filename:
        headers.append(
            (
                "Content-Disposition",
                f"attachment; filename*=UTF-8''{quote(filename)}",
            )
        )
    return Response(body=content, headers=headers)


def wants_ajax(request: Request) -> bool:
    return request.environ.get("HTTP_X_REQUESTED_WITH", "").lower() == "xmlhttprequest"


def session_cookie_header(session_id: str) -> tuple[str, str]:
    cookie = SimpleCookie()
    cookie[SESSION_COOKIE] = session_id
    cookie[SESSION_COOKIE]["path"] = "/"
    cookie[SESSION_COOKIE]["httponly"] = True
    cookie[SESSION_COOKIE]["max-age"] = str(SESSION_HOURS * 3600)
    return ("Set-Cookie", cookie.output(header="").strip())


def clear_session_cookie_header() -> tuple[str, str]:
    cookie = SimpleCookie()
    cookie[SESSION_COOKIE] = ""
    cookie[SESSION_COOKIE]["path"] = "/"
    cookie[SESSION_COOKIE]["httponly"] = True
    cookie[SESSION_COOKIE]["max-age"] = "0"
    return ("Set-Cookie", cookie.output(header="").strip())


def render_template(request: Request, template_name: str, **context: Any) -> Response:
    shell_overrides = context.pop("shell_overrides", None)
    template = env.get_template(template_name)
    base_context = {
        "request": request,
        "flash": request.flash,
        "program_status_labels": PROGRAM_STATUS_LABELS,
        "submission_status_labels": SUBMISSION_STATUS_LABELS,
        "status_label": status_label,
        "format_datetime": format_datetime,
        "shell": build_shell_context(request, template_name, shell_overrides=shell_overrides),
    }
    html = template.render(**base_context, **context)
    return html_response(html)


def require_role(request: Request, role: str) -> Response | None:
    if not request.session or request.session["role"] != role:
        if role == "admin":
            return redirect_response("/admin/login")
        return redirect_response(f"/?role={role}")
    return None


def get_current_admin(request: Request) -> sqlite3.Row | None:
    if not request.session or request.session["role"] != "admin":
        return None
    return request.db.execute(
        "SELECT * FROM admins WHERE username = ?",
        (request.session["admin_username"],),
    ).fetchone()


def get_current_teacher(request: Request) -> sqlite3.Row | None:
    if not request.session or request.session["role"] != "teacher":
        return None
    return request.db.execute(
        "SELECT * FROM teachers WHERE id = ?",
        (request.session["teacher_id"],),
    ).fetchone()


def build_teacher_nav_groups(program_id: int | None = None) -> list[dict[str, Any]]:
    program_status_href = f"/teacher/programs/{program_id}" if program_id else "/teacher"
    review_list_href = f"/teacher/programs/{program_id}/reviews" if program_id else "/teacher"
    return [
        {
            "label": "강사",
            "items": [
                {"key": "teacher-dashboard", "label": "프로그램 리스트", "href": "/teacher"},
                {"key": "teacher-programs", "label": "학생 면담지 제출현황", "href": program_status_href},
                {"key": "teacher-reviews", "label": "개별 면담지 검토", "href": review_list_href},
            ],
        }
    ]


def build_shell_context(
    request: Request,
    template_name: str,
    shell_overrides: dict[str, Any] | None = None,
) -> dict[str, Any]:
    template_meta = {
        "landing.html": {
            "authenticated": False,
            "page_title": "AFE 캠프 산출물 관리 시스템",
            "page_description": "학생, 강사, 관리자가 같은 포털에서 역할을 선택해 로그인합니다.",
            "active_key": "login",
        },
        "admin_login.html": {
            "authenticated": False,
            "page_title": "관리자 로그인",
            "page_description": "관리자 계정으로 운영 대시보드에 접속합니다.",
            "active_key": "login",
        },
        "admin_dashboard.html": {
            "authenticated": True,
            "page_title": "관리자 운영 대시보드",
            "page_description": "캠프, 강사, 결과물, 프로그램 유형을 한 화면에서 운영합니다.",
            "active_key": "admin-dashboard",
        },
        "admin_program_detail.html": {
            "authenticated": True,
            "page_title": "캠프 상세 / 결과물 관리",
            "page_description": "학생 제출, 강사 평가, 관리자 최종 평가를 검토합니다.",
            "active_key": "camp-list",
        },
        "teacher_dashboard.html": {
            "authenticated": True,
            "page_title": "강사 메인 / 프로그램 리스트",
            "page_description": "배정된 캠프와 진행 현황을 확인합니다.",
            "active_key": "teacher-dashboard",
        },
        "teacher_program_detail.html": {
            "authenticated": True,
            "page_title": "학생 면담지 제출현황",
            "page_description": "프로그램 요약과 학생 제출 현황을 한 화면에서 확인합니다.",
            "active_key": "teacher-programs",
        },
        "teacher_submission_list.html": {
            "authenticated": True,
            "page_title": "개별 면담지 리스트",
            "page_description": "학생별 제출 상태를 표로 확인하고 개별 면담지 검토 화면으로 이동합니다.",
            "active_key": "teacher-reviews",
        },
        "teacher_submission_detail.html": {
            "authenticated": True,
            "page_title": "개별 면담지 검토",
            "page_description": "학생 1명의 제출 내용을 검토하고 AI 예시를 바탕으로 평가를 작성합니다.",
            "active_key": "teacher-reviews",
        },
        "student_start.html": {
            "authenticated": True,
            "page_title": "참여 프로그램 / 학생 정보 입력",
            "page_description": "프로그램 정보를 확인하고 학생 정보를 먼저 입력합니다.",
            "active_key": "student-info",
        },
        "student_form.html": {
            "authenticated": True,
            "page_title": "자기평가 질문지 / 탐구내용 작성",
            "page_description": "키워드와 탐구 내용을 정리하고 임시저장 또는 제출합니다.",
            "active_key": "student-questionnaire",
        },
        "student_submitted.html": {
            "authenticated": True,
            "page_title": "제출 완료",
            "page_description": "학생 제출이 완료되었으며 강사 검토 단계로 전달됩니다.",
            "active_key": "student-submit",
        },
    }.get(
        template_name,
        {
            "authenticated": bool(request.session),
            "page_title": "AFE 캠프",
            "page_description": "",
            "active_key": "",
        },
    )
    if template_name == "admin_dashboard.html":
        template_meta = {
            **template_meta,
            **ADMIN_PANEL_META[current_admin_panel(request)],
        }
    if shell_overrides:
        template_meta = {
            **template_meta,
            **{key: value for key, value in shell_overrides.items() if key in {"page_title", "page_description", "active_key"}},
        }

    role = request.session["role"] if request.session else ""
    role_label = {"admin": "관리자", "teacher": "강사", "student": "학생"}.get(role, "게스트")
    user_name = ""
    user_detail = ""
    if role == "admin":
        admin = get_current_admin(request)
        if admin:
            user_name = admin["display_name"]
            user_detail = admin["username"]
    elif role == "teacher":
        teacher = get_current_teacher(request)
        if teacher:
            user_name = teacher["name"]
            user_detail = teacher["username"]
    elif role == "student":
        student_name = get_student_name_from_session(request)
        student_number = get_student_number_from_session(request)
        user_name = student_name or "학생 사용자"
        user_detail = student_number

    nav_groups: list[dict[str, Any]] = []
    if role == "admin":
        nav_groups = [
            {
                "label": "캠프",
                "items": [
                    {"key": "camp-create", "label": "개설", "href": admin_panel_path("camp-create")},
                    {"key": "camp-list", "label": "리스트", "href": admin_panel_path("camp-list")},
                ],
            },
            {
                "label": "강사",
                "items": [
                    {"key": "teacher-create", "label": "등록", "href": admin_panel_path("teacher-create")},
                    {"key": "teacher-list", "label": "리스트", "href": admin_panel_path("teacher-list")},
                    {"key": "teacher-irregular", "label": "배정현황", "href": admin_panel_path("teacher-irregular")},
                ],
            },
            {
                "label": "유형관리",
                "items": [
                    {"key": "template-create", "label": "유형입력", "href": admin_panel_path("template-create")},
                    {"key": "template-manage", "label": "관리", "href": admin_panel_path("template-manage")},
                    {"key": "prompt-manage", "label": "프롬프트 관리", "href": admin_panel_path("prompt-manage")},
                ],
            },
        ]
    elif role == "teacher":
        nav_groups = build_teacher_nav_groups()
    elif role == "student":
        nav_groups = [
            {
                "label": "학생",
                "items": [
                    {"key": "student-info", "label": "프로그램 / 학생 정보", "href": "/student/start"},
                    {"key": "student-questionnaire", "label": "자기평가 질문지", "href": "/student"},
                    {"key": "student-submit", "label": "제출 확인", "href": "/student"},
                ],
            }
        ]

    active_key = template_meta["active_key"]
    if shell_overrides and shell_overrides.get("nav_groups"):
        nav_groups = shell_overrides["nav_groups"]
    for group in nav_groups:
        for item in group["items"]:
            item["is_active"] = item["key"] == active_key

    return {
        "authenticated": template_meta["authenticated"],
        "page_title": template_meta["page_title"],
        "page_description": template_meta["page_description"],
        "role": role,
        "role_label": role_label,
        "user_name": user_name,
        "user_detail": user_detail,
        "nav_groups": nav_groups,
    }


def get_program_teacher_mapping(
    db: sqlite3.Connection,
    program_ids: list[int],
) -> dict[int, list[dict[str, Any]]]:
    if not program_ids:
        return {}
    placeholders = ", ".join("?" for _ in program_ids)
    rows = db.execute(
        f"""
        SELECT
            pt.program_id,
            t.id,
            t.name,
            t.email,
            t.access_code,
            t.university,
            t.username,
            t.temporary_password,
            t.memo,
            t.academic_info
        FROM program_teachers pt
        JOIN teachers t ON t.id = pt.teacher_id
        WHERE pt.program_id IN ({placeholders})
        ORDER BY pt.program_id ASC, LOWER(t.name) ASC, LOWER(t.username) ASC
        """,
        program_ids,
    ).fetchall()
    mapping: dict[int, list[dict[str, Any]]] = {}
    for row in rows:
        teacher = dict(row)
        mapping.setdefault(teacher.pop("program_id"), []).append(teacher)
    return mapping


def attach_program_teacher_metadata(
    program: dict[str, Any],
    assigned_teachers: list[dict[str, Any]],
) -> dict[str, Any]:
    names = [teacher["name"] for teacher in assigned_teachers if teacher.get("name")]
    display_names = [
        f"{teacher['name']} ({teacher['username']})"
        if teacher.get("username")
        else teacher["name"]
        for teacher in assigned_teachers
        if teacher.get("name")
    ]
    code_names = [
        f"{teacher['name']} [{teacher['access_code']}]"
        if teacher.get("access_code")
        else teacher["name"]
        for teacher in assigned_teachers
        if teacher.get("name")
    ]
    access_codes = [
        teacher["access_code"] for teacher in assigned_teachers if teacher.get("access_code")
    ]

    primary_teacher_id = program.get("teacher_id")
    primary_teacher = next(
        (teacher for teacher in assigned_teachers if teacher["id"] == primary_teacher_id),
        assigned_teachers[0] if assigned_teachers else None,
    )

    program["assigned_teachers"] = assigned_teachers
    program["teacher_count"] = len(assigned_teachers)
    program["teacher_name"] = ", ".join(names)
    program["teacher_names"] = ", ".join(names)
    program["teacher_display_names"] = ", ".join(display_names)
    program["teacher_code_names"] = ", ".join(code_names)
    program["teacher_access_code"] = ", ".join(access_codes)

    if primary_teacher:
        program["teacher_username"] = primary_teacher.get("username", "")
        program["teacher_university"] = primary_teacher.get("university", "")
        program["teacher_academic_info"] = primary_teacher.get("academic_info", "")
    else:
        program["teacher_username"] = ""
        program["teacher_university"] = ""
        program["teacher_academic_info"] = ""
    return program


def attach_program_teacher_metadata_bulk(
    db: sqlite3.Connection,
    rows: list[sqlite3.Row | dict[str, Any]],
) -> list[dict[str, Any]]:
    programs = [dict(row) for row in rows]
    mapping = get_program_teacher_mapping(db, [program["id"] for program in programs])
    for program in programs:
        attach_program_teacher_metadata(program, mapping.get(program["id"], []))
    return programs


def get_program_with_teacher(db: sqlite3.Connection, program_id: int) -> dict[str, Any] | None:
    row = db.execute(
        """
        SELECT
            p.*,
            (
                SELECT COUNT(*) FROM submissions s
                WHERE s.program_id = p.id
            ) AS submission_count
        FROM programs p
        WHERE p.id = ?
        """,
        (program_id,),
    ).fetchone()
    if not row:
        return None
    return attach_program_teacher_metadata_bulk(db, [row])[0]


def teacher_has_program_access(db: sqlite3.Connection, teacher_id: int, program_id: int) -> bool:
    row = db.execute(
        "SELECT 1 FROM program_teachers WHERE teacher_id = ? AND program_id = ?",
        (teacher_id, program_id),
    ).fetchone()
    return bool(row)


def get_program_form_schema(program: sqlite3.Row | dict[str, Any]) -> dict[str, Any]:
    raw_schema = parse_json(program["questions_json"], [])
    program_keys = set(program.keys()) if hasattr(program, "keys") else set(program)
    if "template_name" in program_keys:
        fallback_title = program["template_name"]
    elif "title" in program_keys:
        fallback_title = program["title"]
    else:
        fallback_title = ""
    return normalize_template_schema(raw_schema, fallback_title=fallback_title)


def get_program_questions(program: sqlite3.Row | dict[str, Any]) -> list[str]:
    return [field["label"] for field in get_program_form_schema(program)["flat_fields"]]


def get_submissions_for_program(db: sqlite3.Connection, program_id: int) -> list[dict[str, Any]]:
    rows = db.execute(
        """
        SELECT * FROM submissions
        WHERE program_id = ?
        ORDER BY student_number ASC, student_name ASC
        """,
        (program_id,),
    ).fetchall()
    submissions = []
    for row in rows:
        item = dict(row)
        answers = parse_json(row["answers_json"], [])
        normalized_answers = []
        for index, answer in enumerate(answers, start=1):
            if isinstance(answer, dict):
                normalized_answers.append(
                    {
                        "field_id": answer.get("field_id") or f"answer_{index - 1}",
                        "question": answer.get("question") or answer.get("label") or f"문항 {index}",
                        "answer": answer.get("answer", ""),
                        "section_title": answer.get("section_title", ""),
                    }
                )
            else:
                normalized_answers.append(
                    {
                        "field_id": f"answer_{index - 1}",
                        "question": f"문항 {index}",
                        "answer": str(answer),
                        "section_title": "",
                    }
                )
        item["answers"] = normalized_answers
        item["final_evaluation"] = final_evaluation(row)
        submissions.append(item)
    return submissions


def get_answer_map_from_entries(entries: Any) -> dict[str, str]:
    if isinstance(entries, dict):
        return {
            str(key): str(value).strip()
            for key, value in entries.items()
            if str(value).strip()
        }
    answer_map: dict[str, str] = {}
    if isinstance(entries, list):
        for entry in entries:
            if not isinstance(entry, dict):
                continue
            field_id = str(entry.get("field_id") or "").strip()
            answer = str(entry.get("answer") or "").strip()
            if field_id and answer:
                answer_map[field_id] = answer
    return answer_map


def summarize_short_text(value: str, limit: int = 44) -> str:
    text = " ".join(str(value or "").split())
    if not text:
        return "-"
    if len(text) <= limit:
        return text
    return f"{text[: limit - 1].rstrip()}…"


def get_review_status_label(status: str) -> str:
    return {
        "draft": "학생 입력 진행 중",
        "student_submitted": "학생 제출 완료",
        "reviewed": "강사 검토 완료",
        "admin_updated": "강사 검토 완료",
    }.get(status, "학생 입력 진행 중")


def build_submission_review_meta(
    *,
    student_name: str,
    student_number: str,
    desired_major: str,
    status: str,
    answers: Any,
    updated_at: str = "",
    submission_id: int | None = None,
) -> dict[str, Any]:
    answer_map = get_answer_map_from_entries(answers)
    topic = (
        answer_map.get("interest_focus")
        or answer_map.get("curriculum_connection")
        or answer_map.get("highlight_point")
        or answer_map.get("keywords")
        or ""
    )
    career = (
        answer_map.get("career_goal")
        or answer_map.get("career_major")
        or desired_major
        or ""
    )
    return {
        "submission_id": submission_id,
        "student_name": student_name,
        "student_number": student_number,
        "topic": summarize_short_text(topic),
        "career": summarize_short_text(career),
        "status": status,
        "status_label": get_review_status_label(status),
        "updated_at": updated_at,
    }


def get_program_review_rows(db: sqlite3.Connection, program_id: int) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    existing_keys: set[tuple[str, str]] = set()

    for submission in get_submissions_for_program(db, program_id):
        key = (str(submission.get("student_number") or "").strip(), str(submission.get("student_name") or "").strip())
        if key != ("", ""):
            existing_keys.add(key)
        rows.append(
            build_submission_review_meta(
                student_name=submission.get("student_name", ""),
                student_number=submission.get("student_number", ""),
                desired_major=submission.get("desired_major", ""),
                status=submission.get("status", "student_submitted"),
                answers=submission.get("answers", []),
                updated_at=submission.get("teacher_updated_at") or submission.get("student_submitted_at") or "",
                submission_id=int(submission["id"]),
            )
        )

    draft_rows = db.execute(
        """
        SELECT *
        FROM student_drafts
        WHERE program_id = ?
        ORDER BY updated_at DESC, student_number ASC, student_name ASC
        """,
        (program_id,),
    ).fetchall()
    for draft in draft_rows:
        key = (str(draft["student_number"] or "").strip(), str(draft["student_name"] or "").strip())
        if key in existing_keys:
            continue
        rows.append(
            build_submission_review_meta(
                student_name=draft["student_name"],
                student_number=draft["student_number"],
                desired_major=draft["desired_major"] or "",
                status="draft",
                answers=parse_json(draft["answers_json"], {}),
                updated_at=draft["updated_at"] or draft["created_at"] or "",
                submission_id=None,
            )
        )

    rows.sort(
        key=lambda item: (
            item["status"] == "draft",
            item["student_number"] or "999999",
            item["student_name"],
        )
    )
    return rows


def get_student_name_from_session(request: Request) -> str:
    if not request.session or request.session["role"] != "student":
        return ""
    return str((request.session.get("context") or {}).get("student_name", "")).strip()


def get_student_number_from_session(request: Request) -> str:
    if not request.session or request.session["role"] != "student":
        return ""
    return str((request.session.get("context") or {}).get("student_number", "")).strip()


def get_student_desired_major_from_session(request: Request) -> str:
    if not request.session or request.session["role"] != "student":
        return ""
    return str((request.session.get("context") or {}).get("desired_major", "")).strip()


def collect_student_answers(
    request: Request,
    fields: list[dict[str, Any]],
) -> tuple[list[dict[str, str]], dict[str, str]]:
    answers: list[dict[str, str]] = []
    answers_map: dict[str, str] = {}
    for field in fields:
        key = field["id"]
        answer = request.form.get(key, "").strip()
        answers_map[key] = answer
        answers.append(
            {
                "field_id": key,
                "question": field["label"],
                "answer": answer,
                "section_title": field["section_title"],
            }
        )
    return answers, answers_map


def get_student_draft(
    db: sqlite3.Connection,
    program_id: int,
    student_name: str,
    student_number: str,
) -> dict[str, Any] | None:
    if not student_name or not student_number:
        return None
    row = db.execute(
        """
        SELECT * FROM student_drafts
        WHERE program_id = ? AND student_name = ? AND student_number = ?
        """,
        (program_id, student_name, student_number),
    ).fetchone()
    if not row:
        return None
    draft = dict(row)
    parsed_answers = parse_json(draft.get("answers_json"), {})
    draft["answers"] = parsed_answers if isinstance(parsed_answers, dict) else {}
    return draft


def save_student_draft(
    db: sqlite3.Connection,
    *,
    program_id: int,
    student_name: str,
    student_number: str,
    desired_major: str,
    answers_map: dict[str, str],
) -> None:
    if not student_name or not student_number:
        return
    existing = db.execute(
        """
        SELECT id FROM student_drafts
        WHERE program_id = ? AND student_name = ? AND student_number = ?
        """,
        (program_id, student_name, student_number),
    ).fetchone()
    if existing:
        db.execute(
            """
            UPDATE student_drafts
            SET desired_major = ?, answers_json = ?, updated_at = ?
            WHERE id = ?
            """,
            (
                desired_major,
                json_dump(answers_map),
                now_iso(),
                existing["id"],
            ),
        )
    else:
        db.execute(
            """
            INSERT INTO student_drafts (
                program_id, student_name, student_number, desired_major,
                answers_json, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                program_id,
                student_name,
                student_number,
                desired_major,
                json_dump(answers_map),
                now_iso(),
                now_iso(),
            ),
        )
    db.commit()


def delete_student_draft(
    db: sqlite3.Connection,
    *,
    program_id: int,
    student_name: str,
    student_number: str,
) -> None:
    if not student_name or not student_number:
        return
    db.execute(
        "DELETE FROM student_drafts WHERE program_id = ? AND student_name = ? AND student_number = ?",
        (program_id, student_name, student_number),
    )
    db.commit()


def build_student_form_data(
    *,
    student_name: str,
    student_number: str,
    draft: dict[str, Any] | None = None,
    desired_major: str = "",
    answers: dict[str, str] | None = None,
) -> dict[str, Any]:
    draft = draft or {}
    return {
        "student_number": student_number or draft.get("student_number", ""),
        "student_name": student_name,
        "desired_major": desired_major or draft.get("desired_major", ""),
        "answers": answers or draft.get("answers", {}) or {},
    }


def render_student_form_page(
    request: Request,
    *,
    program: dict[str, Any],
    error: str = "",
    is_locked: bool = False,
    form_data: dict[str, Any] | None = None,
) -> Response:
    question_schema = get_program_form_schema(program)
    student_name = (form_data or {}).get("student_name") or get_student_name_from_session(request)
    student_number = (form_data or {}).get("student_number") or get_student_number_from_session(request)
    desired_major = (form_data or {}).get("desired_major") or get_student_desired_major_from_session(request)
    draft = (
        get_student_draft(request.db, int(program["id"]), student_name, student_number)
        if student_name and student_number
        else None
    )
    resolved_form_data = form_data or build_student_form_data(
        student_name=student_name,
        student_number=student_number,
        desired_major=desired_major,
        draft=draft,
    )
    return render_template(
        request,
        "student_form.html",
        program=program,
        questions=get_program_questions(program),
        question_schema=question_schema,
        is_locked=is_locked,
        error=error,
        form_data=resolved_form_data,
        student_name=student_name,
        student_number=student_number,
        has_saved_draft=bool(draft),
        draft_updated_at=draft["updated_at"] if draft else "",
    )


def openai_api_key() -> str:
    return os.environ.get("OPENAI_API_KEY", "").strip()


def openai_is_configured() -> bool:
    return bool(openai_api_key())


def build_ai_example_input(program: sqlite3.Row | dict[str, Any], submission: dict[str, Any]) -> str:
    schema = get_program_form_schema(program)
    lines = [
        "다음 정보를 바탕으로 학교생활기록부용 개별 평가 문장을 작성해 주세요.",
        "",
        "[프로그램 정보]",
        f"- 프로그램명: {program['title']}",
        f"- 프로그램 유형: {program['template_name']}",
        f"- 학교: {program['school_name']}",
        f"- 교급: {program['school_level']}",
        f"- 연도/학기: {program['year']} / {program['semester']}",
        f"- 담당 강사: {program['teacher_name']}",
        "",
        "[학생 기본 정보]",
        f"- 학생명: {submission['student_name']}",
        f"- 학번: {submission['student_number']}",
        f"- 희망전공(학과): {submission['desired_major']}",
    ]
    if schema.get("report_reference"):
        lines.append(f"- 연계 보고서 양식: {schema['report_reference']}")
    lines.extend(["", "[학생 작성 답변]"])
    for answer in submission["answers"]:
        section_title = answer.get("section_title")
        if section_title:
            lines.append(f"- 섹션: {section_title}")
        lines.append(f"  질문: {answer['question']}")
        lines.append(f"  답변: {answer['answer']}")
    lines.extend(["", "위 정보만 사용해서 결과를 작성해 주세요."])
    return "\n".join(lines)


def extract_response_text(payload: dict[str, Any]) -> str:
    parts: list[str] = []
    for output in payload.get("output", []):
        for content in output.get("content", []):
            text_value = content.get("text")
            if isinstance(text_value, dict):
                text_value = text_value.get("value") or text_value.get("text")
            if text_value:
                parts.append(str(text_value))
    if parts:
        return "\n".join(part.strip() for part in parts if part.strip()).strip()
    if payload.get("output_text"):
        return str(payload["output_text"]).strip()
    return ""


def generate_ai_evaluation_example(
    db: sqlite3.Connection,
    program: sqlite3.Row | dict[str, Any],
    submission: dict[str, Any],
) -> tuple[str, str]:
    api_key = openai_api_key()
    if not api_key:
        raise RuntimeError("OPENAI_API_KEY not configured")

    model = os.environ.get("OPENAI_MODEL", "gpt-5-mini").strip() or "gpt-5-mini"
    endpoint = os.environ.get("OPENAI_RESPONSES_URL", "https://api.openai.com/v1/responses").strip()
    reasoning_effort = os.environ.get("OPENAI_REASONING_EFFORT", "low").strip() or "low"
    text_verbosity = os.environ.get("OPENAI_TEXT_VERBOSITY", "low").strip() or "low"
    max_output_tokens = int(os.environ.get("OPENAI_MAX_OUTPUT_TOKENS", "1600"))
    prompt_text = get_teacher_generation_prompt(db, program)
    body = {
        "model": model,
        "reasoning": {"effort": reasoning_effort},
        "text": {
            "format": {"type": "text"},
            "verbosity": text_verbosity,
        },
        "input": [
            {
                "role": "developer",
                "content": [{"type": "input_text", "text": prompt_text}],
            },
            {
                "role": "user",
                "content": [{"type": "input_text", "text": build_ai_example_input(program, submission)}],
            },
        ],
        "max_output_tokens": max_output_tokens,
    }
    request = UrlRequest(
        endpoint,
        data=json_dump(body).encode("utf-8"),
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    timeout_seconds = float(os.environ.get("OPENAI_TIMEOUT_SECONDS", "60"))
    try:
        with urlopen(request, timeout=timeout_seconds) as response:
            payload = json.loads(response.read().decode("utf-8"))
    except HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"OpenAI API error: {detail}") from exc
    except URLError as exc:
        raise RuntimeError(f"OpenAI connection error: {exc}") from exc

    text = extract_response_text(payload)
    if not text:
        incomplete = payload.get("incomplete_details") or {}
        incomplete_reason = incomplete.get("reason")
        if incomplete_reason == "max_output_tokens":
            raise RuntimeError(
                "OpenAI 응답이 토큰 한도에서 중단되었습니다. OPENAI_MAX_OUTPUT_TOKENS 값을 더 크게 설정해 주세요."
            )
        status = payload.get("status") or "unknown"
        raise RuntimeError(f"OpenAI response did not include text output (status={status})")
    return text, model


def ensure_ai_suggestion_for_submission(
    db: sqlite3.Connection,
    program: sqlite3.Row | dict[str, Any],
    submission: dict[str, Any],
    *,
    force: bool = False,
) -> dict[str, Any]:
    existing = (submission.get("ai_suggestion") or "").strip()
    if existing and not force:
        return submission
    if not openai_is_configured():
        submission["ai_error"] = "OPENAI_API_KEY가 설정되지 않았습니다. Render 환경변수에 키를 등록한 뒤 다시 시도해 주세요."
        return submission

    try:
        suggestion, model = generate_ai_evaluation_example(db, program, submission)
    except Exception as exc:
        submission["ai_error"] = str(exc)
        return submission

    generated_at = now_iso()
    db.execute(
        """
        UPDATE submissions
        SET ai_suggestion = ?, ai_generated_at = ?, ai_model = ?
        WHERE id = ?
        """,
        (suggestion, generated_at, model, submission["id"]),
    )
    db.commit()
    submission["ai_suggestion"] = suggestion
    submission["ai_generated_at"] = generated_at
    submission["ai_model"] = model
    submission["ai_error"] = ""
    return submission


def admin_filters_from_request(request: Request) -> dict[str, str]:
    return {
        "year": request.query.get("year", "").strip(),
        "semester": request.query.get("semester", "").strip(),
        "school_name": request.query.get("school_name", "").strip(),
        "teacher_id": request.query.get("teacher_id", "").strip(),
        "status": request.query.get("status", "").strip(),
        "keyword": request.query.get("keyword", "").strip(),
    }


def list_program_options(db: sqlite3.Connection) -> list[dict[str, Any]]:
    rows = db.execute(
        """
        SELECT
            p.id,
            p.title,
            p.program_code,
            p.year,
            p.semester,
            p.school_name,
            p.school_level
        FROM programs p
        ORDER BY p.year DESC, p.semester DESC, p.created_at DESC
        """
    ).fetchall()
    return [dict(row) for row in rows]


def build_teacher_management_rows(
    db: sqlite3.Connection,
    teachers: list[sqlite3.Row | dict[str, Any]],
    program_options: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    assignments = db.execute(
        """
        SELECT
            pt.teacher_id,
            p.id AS program_id,
            p.title AS program_title,
            p.program_code,
            p.year,
            p.semester,
            p.school_name,
            p.school_level
        FROM program_teachers pt
        JOIN programs p ON p.id = pt.program_id
        ORDER BY p.year DESC, p.semester DESC, p.created_at DESC
        """
    ).fetchall()
    assignment_map: dict[int, list[dict[str, Any]]] = {}
    for row in assignments:
        assignment_map.setdefault(row["teacher_id"], []).append(dict(row))

    rows: list[dict[str, Any]] = []
    for teacher in teachers:
        teacher_item = dict(teacher)
        assigned_programs = assignment_map.get(teacher_item["id"], [])
        assigned_ids = {program["program_id"] for program in assigned_programs}
        teacher_item["assigned_programs"] = assigned_programs
        teacher_item["available_programs"] = [
            program for program in program_options if program["id"] not in assigned_ids
        ]
        rows.append(teacher_item)
    return rows


def query_teacher_assignments(db: sqlite3.Connection) -> list[sqlite3.Row]:
    return db.execute(
        """
        SELECT
            t.id AS teacher_id,
            t.name AS teacher_name,
            t.username AS teacher_username,
            t.access_code AS teacher_access_code,
            t.university AS teacher_university,
            t.school_name AS teacher_school_name,
            t.department_name AS teacher_department_name,
            t.academic_info AS teacher_academic_info,
            p.id AS program_id,
            p.title AS program_title,
            p.year,
            p.semester,
            p.school_name,
            p.school_level,
            p.program_code,
            p.status,
            (
                SELECT COUNT(*) FROM submissions s WHERE s.program_id = p.id
            ) AS submission_count
        FROM teachers t
        LEFT JOIN program_teachers pt ON pt.teacher_id = t.id
        LEFT JOIN programs p ON p.id = pt.program_id
        ORDER BY t.created_at DESC, p.year DESC, p.semester DESC, p.created_at DESC
        """
    ).fetchall()


def query_programs(db: sqlite3.Connection, filters: dict[str, str]) -> list[dict[str, Any]]:
    sql = """
        SELECT
            p.*,
            (
                SELECT COUNT(*) FROM submissions s WHERE s.program_id = p.id
            ) AS submission_count,
            (
                SELECT COUNT(*) FROM submissions s
                WHERE s.program_id = p.id
                AND TRIM(COALESCE(s.teacher_evaluation, '')) <> ''
            ) AS reviewed_count
        FROM programs p
        WHERE 1 = 1
    """
    params: list[Any] = []
    if filters["year"]:
        sql += " AND p.year = ?"
        params.append(filters["year"])
    if filters["semester"]:
        sql += " AND p.semester = ?"
        params.append(filters["semester"])
    if filters["school_name"]:
        sql += " AND p.school_name LIKE ?"
        params.append(f"%{filters['school_name']}%")
    if filters["teacher_id"]:
        sql += " AND EXISTS (SELECT 1 FROM program_teachers pt WHERE pt.program_id = p.id AND pt.teacher_id = ?)"
        params.append(filters["teacher_id"])
    if filters["status"]:
        sql += " AND p.status = ?"
        params.append(filters["status"])
    if filters["keyword"]:
        sql += " AND (p.title LIKE ? OR p.program_code LIKE ? OR p.template_name LIKE ?)"
        params.extend([f"%{filters['keyword']}%"] * 3)
    sql += " ORDER BY p.year DESC, p.semester DESC, p.created_at DESC"
    return attach_program_teacher_metadata_bulk(db, db.execute(sql, params).fetchall())


def dashboard_metrics(db: sqlite3.Connection) -> dict[str, int]:
    total_programs = db.execute("SELECT COUNT(*) AS count FROM programs").fetchone()["count"]
    open_programs = db.execute(
        "SELECT COUNT(*) AS count FROM programs WHERE status = 'collecting'"
    ).fetchone()["count"]
    teacher_submitted = db.execute(
        "SELECT COUNT(*) AS count FROM programs WHERE status = 'teacher_submitted'"
    ).fetchone()["count"]
    total_submissions = db.execute(
        "SELECT COUNT(*) AS count FROM submissions"
    ).fetchone()["count"]
    return {
        "total_programs": total_programs,
        "open_programs": open_programs,
        "teacher_submitted": teacher_submitted,
        "total_submissions": total_submissions,
    }


def build_excel(programs: list[sqlite3.Row | dict[str, Any]], db: sqlite3.Connection) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "평가 결과"
    headers = [
        "학년도",
        "학기",
        "학교명",
        "교급",
        "프로그램명",
        "프로그램유형",
        "강사명",
        "프로그램코드",
        "학번",
        "이름",
        "희망전공",
        "평가 내용",
        "상태",
        "학생 제출 시각",
        "강사 수정 시각",
        "관리자 수정 시각",
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for program in programs:
        submissions = get_submissions_for_program(db, program["id"])
        if not submissions:
            sheet.append(
                [
                    program["year"],
                    program["semester"],
                    program["school_name"],
                    program["school_level"],
                    program["title"],
                    program["template_name"],
                    program["teacher_name"],
                    program["program_code"],
                    "",
                    "",
                    "",
                    "",
                    status_label(program["status"], "program"),
                    "",
                    "",
                    "",
                ]
            )
            continue

        for submission in submissions:
            sheet.append(
                [
                    program["year"],
                    program["semester"],
                    program["school_name"],
                    program["school_level"],
                    program["title"],
                    program["template_name"],
                    program["teacher_name"],
                    program["program_code"],
                    submission["student_number"],
                    submission["student_name"],
                    submission["desired_major"],
                    submission["final_evaluation"],
                    status_label(submission["status"], "submission"),
                    format_datetime(submission["student_submitted_at"]),
                    format_datetime(submission["teacher_updated_at"]),
                    format_datetime(submission["admin_updated_at"]),
                ]
            )

    widths = {
        "A": 12,
        "B": 10,
        "C": 20,
        "D": 10,
        "E": 28,
        "F": 14,
        "G": 14,
        "H": 12,
        "I": 10,
        "J": 12,
        "K": 18,
        "L": 52,
        "M": 16,
        "N": 18,
        "O": 18,
        "P": 18,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    info_sheet = workbook.create_sheet("사용 안내")
    info_sheet["A1"] = "AFE 캠프 산출물 관리 시스템"
    info_sheet["A2"] = "다운로드 시각"
    info_sheet["B2"] = format_datetime(now_iso())
    info_sheet["A4"] = "평가 내용은 관리자 최종 평가가 있으면 그 값을 우선 사용하고, 없으면 강사 평가를 사용합니다."
    info_sheet.column_dimensions["A"].width = 70
    info_sheet.column_dimensions["B"].width = 24

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def reference_proxy_response(path: str) -> Response:
    try:
        payload = fetch_reference_api_payload(path)
    except HTTPError as exc:
        return json_response(
            {
                "status": "Error",
                "message": f"Reference API request failed ({exc.code})",
                "data": [],
            },
            status="502 Bad Gateway",
        )
    except (URLError, TimeoutError, ValueError) as exc:
        return json_response(
            {
                "status": "Error",
                "message": f"Reference API is unavailable: {exc}",
                "data": [],
            },
            status="502 Bad Gateway",
        )
    return json_response(payload)


def serve_static(path: str) -> Response:
    relative = path.removeprefix("/static/").strip("/")
    candidate = (STATIC_DIR / relative).resolve()
    if not str(candidate).startswith(str(STATIC_DIR.resolve())) or not candidate.exists():
        return text_response("Not Found", status="404 Not Found")
    content_type, _ = mimetypes.guess_type(candidate.name)
    return Response(
        body=candidate.read_bytes(),
        status="200 OK",
        headers=[("Content-Type", content_type or "application/octet-stream")],
    )


@route("GET", r"/references/cities")
def reference_cities(_: Request) -> Response:
    return reference_proxy_response("/api/references/cities")


@route("GET", r"/references/(?P<city_id>\d+)/districts")
def reference_districts(_: Request, city_id: str) -> Response:
    return reference_proxy_response(f"/api/references/{city_id}/districts")


@route("GET", r"/references/(?P<district_id>\d+)/schools")
def reference_schools(request: Request, district_id: str) -> Response:
    school_level = request.query.get("schoolLevel", "").strip().upper()
    if school_level not in {"MIDDLE", "HIGH"}:
        school_level = "HIGH"
    return reference_proxy_response(
        f"/api/references/{district_id}/schools?schoolLevel={school_level}"
    )


@route("GET", r"/references/curricula")
def reference_curricula(_: Request) -> Response:
    return reference_proxy_response("/api/references/curricula")


@route("GET", r"/references/curricula/(?P<curriculum_id>\d+)")
def reference_curricula_units(_: Request, curriculum_id: str) -> Response:
    return reference_proxy_response(f"/api/references/curricula/{curriculum_id}")


@route("GET", r"/references/curriculumUnit/(?P<curriculum_unit_id>\d+)")
def reference_curricula_sub_units(_: Request, curriculum_unit_id: str) -> Response:
    return reference_proxy_response(f"/api/references/curriculumUnit/{curriculum_unit_id}")


@route("GET", r"/references/desired-careers")
def reference_desired_careers(_: Request) -> Response:
    return reference_proxy_response("/api/references/desired-careers")


@route("GET", r"/")
def landing(request: Request) -> Response:
    active_role = request.query.get("role", "").strip()
    if not active_role and request.session and request.session["role"] in {"admin", "student", "teacher"}:
        active_role = request.session["role"]
    if not active_role:
        active_role = "student"
    if active_role not in {"admin", "student", "teacher"}:
        active_role = "student"
    error = request.query.get("error", "")
    current_teacher = get_current_teacher(request)
    current_program = None
    if request.session and request.session["role"] == "student":
        current_program = get_program_with_teacher(request.db, request.session["program_id"])
    return render_template(
        request,
        "landing.html",
        active_role=active_role,
        error=error,
        current_teacher=current_teacher,
        current_program=current_program,
    )


@route("GET", r"/admin/login")
def admin_login_page(request: Request) -> Response:
    if request.session and request.session["role"] == "admin":
        return redirect_response("/admin")
    error = request.query.get("error", "")
    destination = "/?role=admin"
    if error:
        destination = f"{destination}&error={quote(error)}"
    return redirect_response(destination)


@route("GET", r"/login/admin")
def legacy_admin_login_page(request: Request) -> Response:
    return redirect_response("/admin/login")


@route("POST", r"/login/admin")
def login_admin(request: Request) -> Response:
    username = request.form.get("username", "").strip()
    password = request.form.get("password", "").strip()
    admin = request.db.execute(
        "SELECT * FROM admins WHERE username = ?",
        (username,),
    ).fetchone()
    if not admin or not verify_password(password, admin["password_hash"]):
        return login_error_redirect("admin")

    if request.session:
        destroy_session(request.db, request.session["id"])
    session_id = create_session(request.db, role="admin", admin_username=admin["username"])
    headers = [session_cookie_header(session_id)]
    return redirect_response("/admin", headers=headers)


@route("POST", r"/login/teacher")
def login_teacher(request: Request) -> Response:
    username = request.form.get("username", "").strip().lower()
    password = request.form.get("password", "").strip()
    teacher = None

    if username and password:
        candidate = request.db.execute(
            "SELECT * FROM teachers WHERE username = ?",
            (username,),
        ).fetchone()
        if candidate and verify_password(password, candidate["password_hash"]):
            teacher = candidate

    if not teacher:
        return login_error_redirect("teacher")

    if request.session:
        destroy_session(request.db, request.session["id"])
    session_id = create_session(request.db, role="teacher", teacher_id=teacher["id"])
    headers = [session_cookie_header(session_id)]
    return redirect_response("/teacher", headers=headers)


@route("POST", r"/login/student")
def login_student(request: Request) -> Response:
    program_code = re.sub(r"[^A-Za-z0-9]", "", request.form.get("program_code", "").strip()).upper()[:12]
    program_row = request.db.execute(
        "SELECT * FROM programs WHERE program_code = ?",
        (program_code,),
    ).fetchone()
    program = attach_program_teacher_metadata_bulk(request.db, [program_row])[0] if program_row else None
    if not program:
        return redirect_response("/?role=student&error=유효한%20프로그램%20코드를%20입력해%20주세요.")

    if request.session:
        destroy_session(request.db, request.session["id"])
    session_id = create_session(
        request.db,
        role="student",
        program_id=program["id"],
        context={"program_code": program_code},
    )
    headers = [session_cookie_header(session_id)]
    return redirect_response("/student/start", headers=headers)


@route("GET", r"/logout")
def logout(request: Request) -> Response:
    if request.session:
        destroy_session(request.db, request.session["id"])
    return redirect_response("/", headers=[clear_session_cookie_header()])


@route("GET", r"/admin")
def admin_dashboard(request: Request) -> Response:
    auth = require_role(request, "admin")
    if auth:
        return auth
    admin = get_current_admin(request)
    active_admin_panel = current_admin_panel(request)
    teachers = request.db.execute("SELECT * FROM teachers ORDER BY created_at DESC").fetchall()
    program_options = list_program_options(request.db)
    teacher_management_rows = build_teacher_management_rows(request.db, teachers, program_options)
    teacher_assignments = query_teacher_assignments(request.db)
    template_rows = request.db.execute(
        """
        SELECT
            pt.*,
            (
                SELECT COUNT(*) FROM programs p
                WHERE p.template_id = pt.id
            ) AS usage_count
        FROM program_templates pt
        ORDER BY pt.created_at DESC
        """
    ).fetchall()
    templates = [get_template_card(row) for row in template_rows]
    blank_filters = {
        "year": "",
        "semester": "",
        "school_name": "",
        "teacher_id": "",
        "status": "",
        "keyword": "",
    }
    camp_programs = query_programs(request.db, blank_filters)
    return render_template(
        request,
        "admin_dashboard.html",
        admin=admin,
        active_admin_panel=active_admin_panel,
        teachers=teachers,
        teacher_management_rows=teacher_management_rows,
        program_options=program_options,
        teacher_assignments=teacher_assignments,
        templates=templates,
        camp_programs=camp_programs,
        school_levels=SCHOOL_LEVEL_OPTIONS,
        semesters=SEMESTER_OPTIONS,
        default_prompt_text=EVALUATION_EXAMPLE_PROMPT.strip(),
        teacher_generation_prompt=get_teacher_generation_prompt(request.db),
    )


@route("POST", r"/admin/teachers")
def admin_create_teacher(request: Request) -> Response:
    auth = require_role(request, "admin")
    if auth:
        return auth
    name = request.form.get("name", "").strip()
    email = request.form.get("email", "").strip()
    university = request.form.get("university", "").strip()
    school_name = request.form.get("school_name", "").strip()
    department_name = request.form.get("department_name", "").strip()
    username = request.form.get("username", "").strip().lower()
    raw_password = request.form.get("password", "").strip()
    memo = request.form.get("memo", "").strip()
    academic_info = request.form.get("academic_info", "").strip()
    access_code = request.form.get("access_code", "").strip().upper()
    if not name:
        set_flash(request.db, request.session["id"], "강사 이름을 입력해 주세요.", "error")
        return redirect_response(admin_panel_path("teacher-create"))
    if not username:
        username = generate_teacher_username(request.db, name)
    if request.db.execute("SELECT 1 FROM teachers WHERE username = ?", (username,)).fetchone():
        set_flash(request.db, request.session["id"], "이미 사용 중인 강사 아이디입니다.", "error")
        return redirect_response(admin_panel_path("teacher-create"))
    if not raw_password:
        raw_password = generate_teacher_password()
    if not access_code:
        access_code = generate_teacher_code(request.db)
    try:
        request.db.execute(
            """
            INSERT INTO teachers (
                name, email, access_code, university, school_name, department_name,
                username, password_hash, temporary_password, memo, academic_info, created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                name,
                email,
                access_code,
                university,
                school_name or university,
                department_name,
                username,
                password_hash(raw_password),
                raw_password,
                memo,
                academic_info,
                now_iso(),
            ),
        )
        request.db.commit()
    except sqlite3.IntegrityError:
        set_flash(request.db, request.session["id"], "이미 사용 중인 강사 코드 또는 계정 정보입니다.", "error")
        return redirect_response(admin_panel_path("teacher-create"))
    set_flash(
        request.db,
        request.session["id"],
        f"{name} 강사가 등록되었습니다. 아이디 {username} / 비밀번호 {raw_password} / 코드 {access_code}",
        "success",
    )
    return redirect_response(admin_panel_path("teacher-create"))


@route("POST", r"/admin/teachers/(?P<teacher_id>\d+)/assign")
def admin_assign_teacher_to_program(request: Request, teacher_id: str) -> Response:
    auth = require_role(request, "admin")
    if auth:
        return auth

    program_id = request.form.get("program_id", "").strip()
    teacher = request.db.execute(
        "SELECT id, name FROM teachers WHERE id = ?",
        (teacher_id,),
    ).fetchone()
    program = request.db.execute(
        "SELECT id, title, program_code FROM programs WHERE id = ?",
        (program_id,),
    ).fetchone()
    if not teacher or not program:
        set_flash(request.db, request.session["id"], "배정할 강사 또는 캠프 정보를 찾을 수 없습니다.", "error")
        return redirect_response(admin_panel_path("teacher-list"))

    exists = request.db.execute(
        "SELECT 1 FROM program_teachers WHERE teacher_id = ? AND program_id = ?",
        (teacher_id, program_id),
    ).fetchone()
    if exists:
        set_flash(request.db, request.session["id"], "이미 배정된 캠프입니다.", "info")
        return redirect_response(admin_panel_path("teacher-list"))

    request.db.execute(
        """
        INSERT OR IGNORE INTO program_teachers (program_id, teacher_id, assigned_at)
        VALUES (?, ?, ?)
        """,
        (program_id, teacher_id, now_iso()),
    )
    request.db.commit()
    set_flash(
        request.db,
        request.session["id"],
        f"{teacher['name']} 강사를 {program['title']} ({program['program_code']}) 캠프에 배정했습니다.",
        "success",
    )
    return redirect_response(admin_panel_path("teacher-list"))


@route("POST", r"/admin/teachers/(?P<teacher_id>\d+)/programs/(?P<program_id>\d+)/remove")
def admin_remove_teacher_from_program(request: Request, teacher_id: str, program_id: str) -> Response:
    auth = require_role(request, "admin")
    if auth:
        return auth

    teacher = request.db.execute(
        "SELECT id, name FROM teachers WHERE id = ?",
        (teacher_id,),
    ).fetchone()
    program = request.db.execute(
        "SELECT id, title, program_code, teacher_id FROM programs WHERE id = ?",
        (program_id,),
    ).fetchone()
    assignment = request.db.execute(
        "SELECT 1 FROM program_teachers WHERE teacher_id = ? AND program_id = ?",
        (teacher_id, program_id),
    ).fetchone()
    if not teacher or not program or not assignment:
        set_flash(request.db, request.session["id"], "배정 취소할 정보를 찾을 수 없습니다.", "error")
        return redirect_response(admin_panel_path("teacher-list"))

    assigned_count_row = request.db.execute(
        "SELECT COUNT(*) AS count FROM program_teachers WHERE program_id = ?",
        (program_id,),
    ).fetchone()
    assigned_count = int(assigned_count_row["count"] if assigned_count_row else 0)
    if assigned_count <= 1:
        set_flash(
            request.db,
            request.session["id"],
            "캠프에는 최소 한 명의 강사가 배정되어 있어야 하므로 마지막 강사는 배정 취소할 수 없습니다.",
            "error",
        )
        return redirect_response(admin_panel_path("teacher-list"))

    next_teacher = request.db.execute(
        """
        SELECT teacher_id
        FROM program_teachers
        WHERE program_id = ? AND teacher_id != ?
        ORDER BY assigned_at ASC
        LIMIT 1
        """,
        (program_id, teacher_id),
    ).fetchone()

    request.db.execute(
        "DELETE FROM program_teachers WHERE teacher_id = ? AND program_id = ?",
        (teacher_id, program_id),
    )
    if int(program["teacher_id"]) == int(teacher_id) and next_teacher:
        request.db.execute(
            "UPDATE programs SET teacher_id = ? WHERE id = ?",
            (next_teacher["teacher_id"], program_id),
        )
    request.db.commit()
    set_flash(
        request.db,
        request.session["id"],
        f"{teacher['name']} 강사의 {program['title']} ({program['program_code']}) 배정을 취소했습니다.",
        "success",
    )
    return redirect_response(admin_panel_path("teacher-list"))


@route("POST", r"/admin/templates")
def admin_create_template(request: Request) -> Response:
    auth = require_role(request, "admin")
    if auth:
        return auth
    name = request.form.get("name", "").strip()
    description = request.form.get("description", "").strip()
    prompt_text = request.form.get("prompt_text", "").strip() or EVALUATION_EXAMPLE_PROMPT
    questions = parse_questions_from_text(request.form.get("questions", ""))
    if not name or not questions:
        set_flash(request.db, request.session["id"], "프로그램 유형명과 질문 목록을 모두 입력해 주세요.", "error")
        return redirect_response(admin_panel_path("template-create"))
    try:
        request.db.execute(
            """
            INSERT INTO program_templates (name, description, prompt_text, questions_json, created_at)
            VALUES (?, ?, ?, ?, ?)
            """,
            (name, description, prompt_text, json_dump(questions), now_iso()),
        )
        request.db.commit()
    except sqlite3.IntegrityError:
        set_flash(request.db, request.session["id"], "같은 이름의 프로그램 유형이 이미 존재합니다.", "error")
        return redirect_response(admin_panel_path("template-create"))
    set_flash(request.db, request.session["id"], f"{name} 유형이 추가되었습니다.", "success")
    return redirect_response(admin_panel_path("template-create"))


@route("POST", r"/admin/templates/(?P<template_id>\d+)")
def admin_update_template(request: Request, template_id: str) -> Response:
    auth = require_role(request, "admin")
    if auth:
        return auth
    name = request.form.get("name", "").strip()
    description = request.form.get("description", "").strip()
    prompt_text = request.form.get("prompt_text", "").strip() or EVALUATION_EXAMPLE_PROMPT
    questions = parse_questions_from_text(request.form.get("questions", ""))
    if not name or not questions:
        set_flash(request.db, request.session["id"], "유형명, 프롬프트, 질문 내용을 확인해 주세요.", "error")
        return redirect_response(admin_panel_path("template-manage"))

    exists = request.db.execute(
        "SELECT 1 FROM program_templates WHERE name = ? AND id != ?",
        (name, template_id),
    ).fetchone()
    if exists:
        set_flash(request.db, request.session["id"], "같은 이름의 프로그램 유형이 이미 존재합니다.", "error")
        return redirect_response(admin_panel_path("template-manage"))

    request.db.execute(
        """
        UPDATE program_templates
        SET name = ?, description = ?, prompt_text = ?, questions_json = ?
        WHERE id = ?
        """,
        (name, description, prompt_text, json_dump(questions), template_id),
    )
    request.db.commit()
    set_flash(request.db, request.session["id"], f"{name} 유형이 수정되었습니다.", "success")
    return redirect_response(admin_panel_path("template-manage"))


@route("POST", r"/admin/templates/(?P<template_id>\d+)/delete")
def admin_delete_template(request: Request, template_id: str) -> Response:
    auth = require_role(request, "admin")
    if auth:
        return auth
    template = request.db.execute(
        "SELECT * FROM program_templates WHERE id = ?",
        (template_id,),
    ).fetchone()
    if not template:
        set_flash(request.db, request.session["id"], "삭제할 프로그램 유형을 찾을 수 없습니다.", "error")
        return redirect_response(admin_panel_path("template-manage"))

    usage_count = request.db.execute(
        "SELECT COUNT(*) AS count FROM programs WHERE template_id = ?",
        (template_id,),
    ).fetchone()["count"]
    if usage_count:
        set_flash(
            request.db,
            request.session["id"],
            f"{template['name']} 유형은 현재 {usage_count}개 프로그램에서 사용 중이라 삭제할 수 없습니다.",
            "error",
        )
        return redirect_response(admin_panel_path("template-manage"))

    request.db.execute("DELETE FROM program_templates WHERE id = ?", (template_id,))
    request.db.commit()
    set_flash(request.db, request.session["id"], f"{template['name']} 유형이 삭제되었습니다.", "success")
    return redirect_response(admin_panel_path("template-manage"))


@route("POST", r"/admin/prompts")
def admin_update_teacher_prompt(request: Request) -> Response:
    auth = require_role(request, "admin")
    if auth:
        return auth
    prompt_text = request.form.get("teacher_generation_prompt", "").strip() or EVALUATION_EXAMPLE_PROMPT.strip()
    set_meta_value(request.db, "teacher_generation_prompt", prompt_text)
    set_flash(request.db, request.session["id"], "강사 생활기록부 생성 프롬프트가 저장되었습니다.", "success")
    return redirect_response(admin_panel_path("prompt-manage"))


@route("POST", r"/admin/programs")
def admin_create_program(request: Request) -> Response:
    auth = require_role(request, "admin")
    if auth:
        return auth

    title = request.form.get("title", "").strip()
    school_name = request.form.get("school_name", "").strip() or "-"
    school_level = request.form.get("school_level", "").strip() or "-"
    year = request.form.get("year", "").strip()
    semester = request.form.get("semester", "").strip()
    template_id = request.form.get("template_id", "").strip()
    teacher_ids_raw = request.form.get("teacher_ids", "").strip()

    teacher_ids: list[int] = []
    for chunk in teacher_ids_raw.split(","):
        value = chunk.strip()
        if value.isdigit():
            parsed = int(value)
            if parsed not in teacher_ids:
                teacher_ids.append(parsed)

    if not all([title, year, semester, template_id]) or not teacher_ids:
        set_flash(request.db, request.session["id"], "프로그램 개설 항목을 모두 입력해 주세요.", "error")
        return redirect_response(admin_panel_path("camp-create"))

    template = request.db.execute(
        "SELECT * FROM program_templates WHERE id = ?",
        (template_id,),
    ).fetchone()
    placeholders = ", ".join("?" for _ in teacher_ids)
    teachers = request.db.execute(
        f"SELECT * FROM teachers WHERE id IN ({placeholders})",
        teacher_ids,
    ).fetchall()
    if not template or len(teachers) != len(teacher_ids):
        set_flash(request.db, request.session["id"], "템플릿 또는 강사 정보가 올바르지 않습니다.", "error")
        return redirect_response(admin_panel_path("camp-create"))

    program_code = generate_program_code(request.db)
    program_params = (
        title,
        school_name,
        school_level,
        int(year),
        semester,
        template["id"],
        template["name"],
        template["description"],
        (template["prompt_text"] or EVALUATION_EXAMPLE_PROMPT),
        template["questions_json"],
        teacher_ids[0],
        program_code,
        "collecting",
        None,
        now_iso(),
    )
    if request.db.dialect == "postgres":
        cursor = request.db.execute(
            """
            INSERT INTO programs (
                title, school_name, school_level, year, semester,
                template_id, template_name, template_description, prompt_text, questions_json,
                teacher_id, program_code, status, teacher_submitted_at, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            RETURNING id
            """,
            program_params,
        )
        program_id = int(cursor.fetchone()["id"])
    else:
        cursor = request.db.execute(
            """
            INSERT INTO programs (
                title, school_name, school_level, year, semester,
                template_id, template_name, template_description, prompt_text, questions_json,
                teacher_id, program_code, status, teacher_submitted_at, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            program_params,
        )
        program_id = int(cursor.lastrowid)
    assigned_at = now_iso()
    for teacher_id in teacher_ids:
        request.db.execute(
            """
            INSERT OR IGNORE INTO program_teachers (program_id, teacher_id, assigned_at)
            VALUES (?, ?, ?)
            """,
            (program_id, teacher_id, assigned_at),
        )
    request.db.commit()
    set_flash(
        request.db,
        request.session["id"],
        f"프로그램이 개설되었습니다. 학생용 프로그램 코드는 {program_code} 입니다. 배정 강사는 {len(teacher_ids)}명입니다.",
        "success",
    )
    return redirect_response(admin_panel_path("camp-create"))


@route("POST", r"/admin/programs/(?P<program_id>\d+)/delete")
def admin_delete_program(request: Request, program_id: str) -> Response:
    auth = require_role(request, "admin")
    if auth:
        return auth
    program = request.db.execute(
        "SELECT id, title, program_code FROM programs WHERE id = ?",
        (program_id,),
    ).fetchone()
    if not program:
        set_flash(request.db, request.session["id"], "삭제할 프로그램을 찾을 수 없습니다.", "error")
        return redirect_response(admin_panel_path("camp-list"))

    request.db.execute("DELETE FROM sessions WHERE program_id = ?", (program_id,))
    request.db.execute("DELETE FROM programs WHERE id = ?", (program_id,))
    request.db.commit()
    set_flash(
        request.db,
        request.session["id"],
        f"{program['title']} 프로그램 ({program['program_code']}) 이 삭제되었습니다.",
        "success",
    )
    if wants_ajax(request):
        return text_response("ok")
    redirect_panel = request.form.get("redirect_panel", "").strip()
    return redirect_response(admin_panel_path(redirect_panel or "camp-list"))


@route("GET", r"/admin/programs/(?P<program_id>\d+)")
def admin_program_detail(request: Request, program_id: str) -> Response:
    auth = require_role(request, "admin")
    if auth:
        return auth
    program = get_program_with_teacher(request.db, int(program_id))
    if not program:
        return text_response("프로그램을 찾을 수 없습니다.", status="404 Not Found")
    submissions = get_submissions_for_program(request.db, int(program_id))
    question_schema = get_program_form_schema(program)
    return render_template(
        request,
        "admin_program_detail.html",
        program=program,
        questions=get_program_questions(program),
        question_schema=question_schema,
        submissions=submissions,
    )


@route("POST", r"/admin/programs/(?P<program_id>\d+)/submissions/(?P<submission_id>\d+)")
def admin_update_submission(request: Request, program_id: str, submission_id: str) -> Response:
    auth = require_role(request, "admin")
    if auth:
        return auth
    admin_feedback = request.form.get("admin_feedback", "").strip()
    request.db.execute(
        """
        UPDATE submissions
        SET admin_feedback = ?, admin_updated_at = ?, status = 'admin_updated'
        WHERE id = ? AND program_id = ?
        """,
        (admin_feedback, now_iso(), submission_id, program_id),
    )
    request.db.commit()
    set_flash(request.db, request.session["id"], "관리자 최종 평가가 저장되었습니다.", "success")
    return redirect_response(f"/admin/programs/{program_id}")


@route("POST", r"/admin/programs/(?P<program_id>\d+)/status")
def admin_update_program_status(request: Request, program_id: str) -> Response:
    auth = require_role(request, "admin")
    if auth:
        return auth
    new_status = request.form.get("status", "").strip()
    if new_status not in {"collecting", "teacher_submitted", "completed"}:
        set_flash(request.db, request.session["id"], "변경할 수 없는 상태값입니다.", "error")
        return redirect_response(f"/admin/programs/{program_id}")
    request.db.execute(
        "UPDATE programs SET status = ? WHERE id = ?",
        (new_status, program_id),
    )
    request.db.commit()
    set_flash(request.db, request.session["id"], "프로그램 상태가 변경되었습니다.", "success")
    return redirect_response(f"/admin/programs/{program_id}")


@route("GET", r"/admin/programs/(?P<program_id>\d+)/download\.xlsx")
def admin_download_program_excel(request: Request, program_id: str) -> Response:
    auth = require_role(request, "admin")
    if auth:
        return auth
    program = get_program_with_teacher(request.db, int(program_id))
    if not program:
        return text_response("프로그램을 찾을 수 없습니다.", status="404 Not Found")
    content = build_excel([program], request.db)
    filename = f"afe_program_{program['id']}.xlsx"
    return bytes_response(
        content,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
    )


@route("GET", r"/admin/download\.xlsx")
def admin_download_filtered_excel(request: Request) -> Response:
    auth = require_role(request, "admin")
    if auth:
        return auth
    filters = admin_filters_from_request(request)
    programs = query_programs(request.db, filters)
    content = build_excel(programs, request.db)
    filename = f"afe_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return bytes_response(
        content,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
    )


@route("GET", r"/teacher")
def teacher_dashboard(request: Request) -> Response:
    auth = require_role(request, "teacher")
    if auth:
        return auth
    teacher = get_current_teacher(request)
    if not teacher:
        return redirect_response("/logout")
    program_rows = request.db.execute(
        """
        SELECT
            p.*,
            (
                SELECT COUNT(*) FROM submissions s WHERE s.program_id = p.id
            ) AS submission_count
        FROM programs p
        WHERE EXISTS (
            SELECT 1 FROM program_teachers pt
            WHERE pt.program_id = p.id AND pt.teacher_id = ?
        )
        ORDER BY p.year DESC, p.semester DESC, p.created_at DESC
        """,
        (teacher["id"],),
    ).fetchall()
    programs = attach_program_teacher_metadata_bulk(request.db, program_rows)
    metrics = {
        "assigned_programs": len(programs),
        "open_programs": len([program for program in programs if program["status"] == "collecting"]),
        "teacher_submitted": len([program for program in programs if program["status"] == "teacher_submitted"]),
        "student_submissions": sum(program["submission_count"] for program in programs),
    }
    spotlight_program = programs[0] if programs else None
    recent_submission_row = request.db.execute(
        """
        SELECT
            s.student_name,
            s.student_number,
            s.student_submitted_at,
            p.id AS program_id,
            p.program_code,
            p.title AS program_title,
            p.school_name,
            p.year,
            p.semester
        FROM submissions s
        JOIN programs p ON p.id = s.program_id
        WHERE EXISTS (
            SELECT 1 FROM program_teachers pt
            WHERE pt.program_id = p.id AND pt.teacher_id = ?
        )
        ORDER BY COALESCE(s.teacher_updated_at, s.student_submitted_at) DESC, s.id DESC
        LIMIT 1
        """,
        (teacher["id"],),
    ).fetchone()
    recent_submission = dict(recent_submission_row) if recent_submission_row else None
    return render_template(
        request,
        "teacher_dashboard.html",
        teacher=teacher,
        programs=programs,
        metrics=metrics,
        spotlight_program=spotlight_program,
        recent_submission=recent_submission,
        shell_overrides={
            "nav_groups": build_teacher_nav_groups(int(spotlight_program["id"])) if spotlight_program else build_teacher_nav_groups(),
            "active_key": "teacher-dashboard",
        },
    )


@route("GET", r"/teacher/programs/(?P<program_id>\d+)")
def teacher_program_detail(request: Request, program_id: str) -> Response:
    auth = require_role(request, "teacher")
    if auth:
        return auth
    teacher = get_current_teacher(request)
    if not teacher:
        return redirect_response("/logout")
    program = get_program_with_teacher(request.db, int(program_id))
    if not program or not teacher_has_program_access(request.db, teacher["id"], int(program_id)):
        return text_response("접근 권한이 없습니다.", status="403 Forbidden")
    return render_template(
        request,
        "teacher_program_detail.html",
        teacher=teacher,
        program=program,
        review_rows=get_program_review_rows(request.db, int(program_id)),
        is_locked=program["status"] in {"teacher_submitted", "completed"},
        shell_overrides={
            "nav_groups": build_teacher_nav_groups(int(program_id)),
            "active_key": "teacher-programs",
        },
    )


@route("GET", r"/teacher/programs/(?P<program_id>\d+)/reviews")
def teacher_submission_list(request: Request, program_id: str) -> Response:
    auth = require_role(request, "teacher")
    if auth:
        return auth
    teacher = get_current_teacher(request)
    if not teacher:
        return redirect_response("/logout")
    program = get_program_with_teacher(request.db, int(program_id))
    if not program or not teacher_has_program_access(request.db, teacher["id"], int(program_id)):
        return text_response("접근 권한이 없습니다.", status="403 Forbidden")
    return render_template(
        request,
        "teacher_submission_list.html",
        teacher=teacher,
        program=program,
        review_rows=get_program_review_rows(request.db, int(program_id)),
        shell_overrides={
            "nav_groups": build_teacher_nav_groups(int(program_id)),
            "active_key": "teacher-reviews",
        },
    )


@route("GET", r"/teacher/programs/(?P<program_id>\d+)/reviews/(?P<submission_id>\d+)")
def teacher_submission_detail(request: Request, program_id: str, submission_id: str) -> Response:
    auth = require_role(request, "teacher")
    if auth:
        return auth
    teacher = get_current_teacher(request)
    if not teacher:
        return redirect_response("/logout")
    program = get_program_with_teacher(request.db, int(program_id))
    if not program or not teacher_has_program_access(request.db, teacher["id"], int(program_id)):
        return text_response("접근 권한이 없습니다.", status="403 Forbidden")
    submissions = get_submissions_for_program(request.db, int(program_id))
    submission = next((item for item in submissions if item["id"] == int(submission_id)), None)
    if not submission:
        return text_response("학생 제출 정보를 찾을 수 없습니다.", status="404 Not Found")
    return render_template(
        request,
        "teacher_submission_detail.html",
        teacher=teacher,
        program=program,
        submission=submission,
        is_locked=program["status"] in {"teacher_submitted", "completed"},
        ai_enabled=openai_is_configured(),
        shell_overrides={
            "nav_groups": build_teacher_nav_groups(int(program_id)),
            "active_key": "teacher-reviews",
        },
    )


@route("POST", r"/teacher/programs/(?P<program_id>\d+)/submissions/(?P<submission_id>\d+)/ai-suggestion")
def teacher_regenerate_ai_suggestion(request: Request, program_id: str, submission_id: str) -> Response:
    auth = require_role(request, "teacher")
    if auth:
        return auth
    teacher = get_current_teacher(request)
    if not teacher:
        return redirect_response("/logout")
    program = get_program_with_teacher(request.db, int(program_id))
    if not program or not teacher_has_program_access(request.db, teacher["id"], int(program_id)):
        return text_response("접근 권한이 없습니다.", status="403 Forbidden")

    row = request.db.execute(
        "SELECT * FROM submissions WHERE id = ? AND program_id = ?",
        (submission_id, program_id),
    ).fetchone()
    if not row:
        set_flash(request.db, request.session["id"], "학생 제출 정보를 찾을 수 없습니다.", "error")
        return redirect_response(f"/teacher/programs/{program_id}/reviews")

    normalized_submission = get_submissions_for_program(request.db, int(program_id))
    target = next((item for item in normalized_submission if item["id"] == int(submission_id)), None)
    if not target:
        set_flash(request.db, request.session["id"], "학생 제출 정보를 찾을 수 없습니다.", "error")
        return redirect_response(f"/teacher/programs/{program_id}/reviews")

    ensure_ai_suggestion_for_submission(request.db, program, target, force=True)
    if target.get("ai_error"):
        set_flash(request.db, request.session["id"], f"평가 내용 예시 생성에 실패했습니다. {target['ai_error']}", "error")
    else:
        set_flash(request.db, request.session["id"], "평가 내용 예시를 다시 생성했습니다.", "success")
    return redirect_response(f"/teacher/programs/{program_id}/reviews/{submission_id}")


@route("POST", r"/teacher/programs/(?P<program_id>\d+)/submissions/(?P<submission_id>\d+)")
def teacher_update_submission(request: Request, program_id: str, submission_id: str) -> Response:
    auth = require_role(request, "teacher")
    if auth:
        return auth
    teacher = get_current_teacher(request)
    if not teacher:
        return redirect_response("/logout")
    program = get_program_with_teacher(request.db, int(program_id))
    if not program or not teacher_has_program_access(request.db, teacher["id"], int(program_id)):
        return text_response("접근 권한이 없습니다.", status="403 Forbidden")
    if program["status"] in {"teacher_submitted", "completed"}:
        set_flash(request.db, request.session["id"], "이미 제출이 완료된 프로그램은 수정할 수 없습니다.", "error")
        return redirect_response(f"/teacher/programs/{program_id}/reviews/{submission_id}")

    teacher_summary = request.form.get("teacher_summary", "").strip()
    teacher_evaluation = request.form.get("teacher_evaluation", "").strip()
    request.db.execute(
        """
        UPDATE submissions
        SET teacher_summary = ?, teacher_evaluation = ?, teacher_updated_at = ?, status = 'reviewed'
        WHERE id = ? AND program_id = ?
        """,
        (teacher_summary, teacher_evaluation, now_iso(), submission_id, program_id),
    )
    request.db.commit()
    set_flash(request.db, request.session["id"], "학생 평가 내용이 저장되었습니다.", "success")
    return redirect_response(f"/teacher/programs/{program_id}/reviews/{submission_id}")


@route("POST", r"/teacher/programs/(?P<program_id>\d+)/submit")
def teacher_submit_program(request: Request, program_id: str) -> Response:
    auth = require_role(request, "teacher")
    if auth:
        return auth
    teacher = get_current_teacher(request)
    if not teacher:
        return redirect_response("/logout")
    program = get_program_with_teacher(request.db, int(program_id))
    if not program or not teacher_has_program_access(request.db, teacher["id"], int(program_id)):
        return text_response("접근 권한이 없습니다.", status="403 Forbidden")
    request.db.execute(
        """
        UPDATE programs
        SET status = 'teacher_submitted', teacher_submitted_at = ?
        WHERE id = ?
        """,
        (now_iso(), program_id),
    )
    request.db.commit()
    set_flash(request.db, request.session["id"], "강사 최종 제출이 완료되었습니다. 이제 관리자가 확인할 수 있습니다.", "success")
    return redirect_response(f"/teacher/programs/{program_id}")


@route("GET", r"/student")
def student_form(request: Request) -> Response:
    auth = require_role(request, "student")
    if auth:
        return auth
    program = get_program_with_teacher(request.db, int(request.session["program_id"]))
    if not program:
        return redirect_response("/?role=student&error=프로그램%20정보를%20찾을%20수%20없습니다.")
    if not get_student_name_from_session(request) or not get_student_number_from_session(request):
        return redirect_response("/student/start")
    return render_student_form_page(
        request,
        program=program,
        is_locked=program["status"] != "collecting",
    )


@route("GET", r"/student/start")
def student_start(request: Request) -> Response:
    auth = require_role(request, "student")
    if auth:
        return auth
    program = get_program_with_teacher(request.db, int(request.session["program_id"]))
    if not program:
        return redirect_response("/?role=student&error=프로그램%20정보를%20찾을%20수%20없습니다.")
    student_name = get_student_name_from_session(request)
    student_number = get_student_number_from_session(request)
    desired_major = get_student_desired_major_from_session(request)
    if student_name and student_number and desired_major and request.query.get("change", "") != "1":
        return redirect_response("/student")
    draft = (
        get_student_draft(request.db, int(program["id"]), student_name, student_number)
        if student_name and student_number
        else None
    )
    return render_template(
        request,
        "student_start.html",
        program=program,
        student_name=student_name,
        student_number=student_number,
        desired_major=desired_major or (draft.get("desired_major", "") if draft else ""),
        draft_exists=bool(draft),
        draft_updated_at=draft["updated_at"] if draft else "",
        error="",
    )


@route("POST", r"/student/start")
def student_start_submit(request: Request) -> Response:
    auth = require_role(request, "student")
    if auth:
        return auth
    program = get_program_with_teacher(request.db, int(request.session["program_id"]))
    if not program:
        return redirect_response("/?role=student&error=프로그램%20정보를%20찾을%20수%20없습니다.")
    student_name = request.form.get("student_name", "").strip()
    student_number = request.form.get("student_number", "").strip()
    desired_major = request.form.get("desired_major", "").strip()
    if not student_name or not student_number or not desired_major:
        draft = (
            get_student_draft(request.db, int(program["id"]), student_name, student_number)
            if student_name and student_number
            else None
        )
        return render_template(
            request,
            "student_start.html",
            program=program,
            student_name=student_name,
            student_number=student_number,
            desired_major=desired_major or (draft.get("desired_major", "") if draft else ""),
            draft_exists=bool(draft),
            draft_updated_at=draft["updated_at"] if draft else "",
            error="학생명, 학번, 희망 전공을 모두 입력해 주세요.",
        )

    updated_context = update_session_context(
        request.db,
        request.session["id"],
        {
            "student_name": student_name,
            "student_number": student_number,
            "desired_major": desired_major,
        },
    )
    if request.session:
        request.session["context"] = updated_context

    draft = get_student_draft(request.db, int(program["id"]), student_name, student_number)
    if draft:
        set_flash(request.db, request.session["id"], "같은 이름과 학번으로 임시저장한 내용을 불러왔습니다.", "info")
    return redirect_response("/student")


@route("POST", r"/student/draft")
def student_save_draft_action(request: Request) -> Response:
    auth = require_role(request, "student")
    if auth:
        return auth
    program = get_program_with_teacher(request.db, int(request.session["program_id"]))
    if not program:
        return redirect_response("/?role=student&error=프로그램%20정보를%20찾을%20수%20없습니다.")
    if program["status"] != "collecting":
        set_flash(request.db, request.session["id"], "지금은 임시저장할 수 없는 프로그램 상태입니다.", "error")
        return redirect_response("/student")

    student_name = get_student_name_from_session(request)
    student_number = get_student_number_from_session(request)
    if not student_name or not student_number:
        return redirect_response("/student/start")

    question_schema = get_program_form_schema(program)
    fields = question_schema["flat_fields"]
    desired_major = request.form.get("desired_major", "").strip() or get_student_desired_major_from_session(request)
    _, answers_map = collect_student_answers(request, fields)
    has_content = bool(
        desired_major
        or any((value or "").strip() for value in answers_map.values())
    )
    if not has_content:
        set_flash(request.db, request.session["id"], "임시저장할 내용을 먼저 입력해 주세요.", "info")
        return redirect_response("/student")

    save_student_draft(
        request.db,
        program_id=int(program["id"]),
        student_name=student_name,
        student_number=student_number,
        desired_major=desired_major,
        answers_map=answers_map,
    )
    set_flash(
        request.db,
        request.session["id"],
        "임시저장되었습니다. 같은 이름과 학번으로 다시 들어오면 이어서 작성할 수 있습니다.",
        "success",
    )
    return redirect_response("/student")


@route("POST", r"/student/submit")
def student_submit(request: Request) -> Response:
    auth = require_role(request, "student")
    if auth:
        return auth
    program = get_program_with_teacher(request.db, int(request.session["program_id"]))
    if not program:
        return redirect_response("/?role=student&error=프로그램%20정보를%20찾을%20수%20없습니다.")
    student_name = get_student_name_from_session(request)
    student_number = get_student_number_from_session(request)
    if not student_name or not student_number:
        return redirect_response("/student/start")
    if program["status"] != "collecting":
        return render_student_form_page(
            request,
            program=program,
            is_locked=True,
            error="이미 강사 제출 또는 관리자 마감이 완료된 프로그램입니다.",
        )

    question_schema = get_program_form_schema(program)
    fields = question_schema["flat_fields"]
    desired_major = request.form.get("desired_major", "").strip() or get_student_desired_major_from_session(request)

    answers, answers_map = collect_student_answers(request, fields)

    has_missing_required = any(
        field.get("required", True) and not answers_map.get(field["id"], "").strip()
        for field in fields
    )
    if not student_number or not student_name or not desired_major or has_missing_required:
        return render_student_form_page(
            request,
            program=program,
            is_locked=False,
            error="학번, 희망전공, 질문 응답을 모두 입력해 주세요.",
            form_data={
                "student_number": student_number,
                "student_name": student_name,
                "desired_major": desired_major,
                "answers": answers_map,
            },
        )

    existing = request.db.execute(
        """
        SELECT id FROM submissions
        WHERE program_id = ? AND student_number = ?
        """,
        (program["id"], student_number),
    ).fetchone()
    saved_submission_id: int | None = None
    if existing:
        request.db.execute(
            """
            UPDATE submissions
            SET student_name = ?, desired_major = ?, answers_json = ?,
                status = 'student_submitted', student_submitted_at = ?,
                teacher_summary = '', teacher_evaluation = '', teacher_updated_at = NULL,
                ai_suggestion = '', ai_generated_at = NULL, ai_model = '',
                admin_feedback = '', admin_updated_at = NULL
            WHERE id = ?
            """,
            (
                student_name,
                desired_major,
                json_dump(answers),
                now_iso(),
                existing["id"],
            ),
        )
        saved_submission_id = int(existing["id"])
    else:
        insert_params = (
            program["id"],
            student_number,
            student_name,
            desired_major,
            json_dump(answers),
            now_iso(),
        )
        if request.db.dialect == "postgres":
            cursor = request.db.execute(
                """
                INSERT INTO submissions (
                    program_id, student_number, student_name, desired_major,
                    answers_json, status, student_submitted_at,
                    teacher_summary, teacher_evaluation, teacher_updated_at,
                    admin_feedback, admin_updated_at
                ) VALUES (?, ?, ?, ?, ?, 'student_submitted', ?, '', '', NULL, '', NULL)
                RETURNING id
                """,
                insert_params,
            )
            saved_submission_id = int(cursor.fetchone()["id"])
        else:
            cursor = request.db.execute(
                """
                INSERT INTO submissions (
                    program_id, student_number, student_name, desired_major,
                    answers_json, status, student_submitted_at,
                    teacher_summary, teacher_evaluation, teacher_updated_at,
                    admin_feedback, admin_updated_at
                ) VALUES (?, ?, ?, ?, ?, 'student_submitted', ?, '', '', NULL, '', NULL)
                """,
                insert_params,
            )
            saved_submission_id = int(cursor.lastrowid)
    request.db.commit()
    delete_student_draft(
        request.db,
        program_id=int(program["id"]),
        student_name=student_name,
        student_number=student_number,
    )
    saved_submission = None
    if saved_submission_id and openai_is_configured():
        normalized_submission = get_submissions_for_program(request.db, int(program["id"]))
        saved_submission = next((item for item in normalized_submission if item["id"] == saved_submission_id), None)
        if saved_submission:
            ensure_ai_suggestion_for_submission(request.db, program, saved_submission, force=True)
    elif saved_submission_id:
        normalized_submission = get_submissions_for_program(request.db, int(program["id"]))
        saved_submission = next((item for item in normalized_submission if item["id"] == saved_submission_id), None)

    return render_template(
        request,
        "student_submitted.html",
        program=program,
        submission=saved_submission,
    )


@route("GET", r"/health")
def health(_: Request) -> Response:
    return text_response("ok")


def application(environ: dict[str, Any], start_response):
    ensure_runtime_ready()
    db = connect_db()
    request = Request(environ, db)
    try:
        if request.path.startswith("/static/"):
            response = serve_static(request.path)
        else:
            response = None
            for method, pattern, handler in ROUTES:
                if method != request.method:
                    continue
                match = pattern.match(request.path)
                if match:
                    response = handler(request, **match.groupdict())
                    break
            if response is None:
                response = text_response("페이지를 찾을 수 없습니다.", status="404 Not Found")
    except Exception as exc:
        response = html_response(
            f"""
            <html lang="ko">
              <head><meta charset="utf-8"><title>오류</title></head>
              <body style="font-family: sans-serif; padding: 32px;">
                <h1>처리 중 오류가 발생했습니다.</h1>
                <pre>{str(exc)}</pre>
                <p><a href="/">첫 페이지로 돌아가기</a></p>
              </body>
            </html>
            """,
            status="500 Internal Server Error",
        )
    finally:
        db.close()

    status, headers, body = response.as_wsgi()
    start_response(status, headers)
    return body


def main() -> None:
    ensure_runtime_ready()
    host = "127.0.0.1"
    port = int(os.environ.get("PORT", "8000"))
    print(f"AFE 서버 실행 중: http://{host}:{port}")
    with make_server(host, port, application) as server:
        server.serve_forever()


if __name__ == "__main__":
    main()
